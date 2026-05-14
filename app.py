#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
SEAE - Sistema de Evaluación de Alternativas Económicas
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog, ttk
import sqlite3
import json
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from fpdf import FPDF
import csv
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
import tkinter as tk
import math

# ================== HOVER ==================
try:
    import mplcursors
    HOVER_AVAILABLE = True
except ImportError:
    HOVER_AVAILABLE = False

# ================== TEMA ==================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class Tema:
    def __init__(self):
        self.bg = "#0F172A"
        self.card = "#1E293B"
        self.accent = "#3B82F6"
        self.success = "#22C55E"
        self.danger = "#EF4444"
        self.warning = "#F59E0B"
        self.text_primary = "#E2E8F0"
        self.text_secondary = "#94A3B8"
        self.border = "#334155"

tema = Tema()

# ================== FUNCIONES FINANCIERAS BASE ==================
def calcular_vpn(inversion, flujos, tasa, años, salvage=0):
    vpn = -inversion
    flujos_desc = []
    for t in range(1, años+1):
        fd = flujos[t-1] / ((1+tasa)**t)
        vpn += fd
        flujos_desc.append(fd)
    vpn += salvage / ((1+tasa)**años)
    return vpn, flujos_desc

def calcular_cae(vpn, tasa, años):
    if tasa == 0:
        return vpn / años if años != 0 else 0
    factor = (tasa * (1+tasa)**años) / ((1+tasa)**años - 1)
    return vpn * factor

def calcular_tir(inversion, flujos, años, salvage=0, estimacion=0.1, max_iter=100, tol=1e-6):
    def f(tasa):
        vpn = -inversion
        for t in range(1, años+1):
            vpn += flujos[t-1] / ((1+tasa)**t)
        vpn += salvage / ((1+tasa)**años)
        return vpn
    def df(tasa):
        delta = 1e-4
        return (f(tasa+delta) - f(tasa-delta)) / (2*delta)
    tasa = estimacion
    for _ in range(max_iter):
        f_val = f(tasa)
        if abs(f_val) < tol:
            return tasa
        df_val = df(tasa)
        if df_val == 0:
            break
        tasa = tasa - f_val / df_val
        if tasa < 0:
            tasa = 0.01
        if tasa > 1:
            tasa = 0.99
    return None

def calcular_payback_preciso(inversion, flujos):
    acum = 0.0
    for i, flujo in enumerate(flujos, start=1):
        if acum + flujo >= inversion:
            fraccion = (inversion - acum) / flujo if flujo != 0 else 0
            return i - 1 + fraccion
        acum += flujo
    return None

def clasificar_riesgo(proyecto):
    flujos = proyecto.get("flujos", [])
    if not flujos:
        return "Alto"
    media = np.mean(flujos) if flujos else 1
    std = np.std(flujos) if len(flujos)>1 else 0
    volatilidad = std/media if media!=0 else 1.0
    tir = proyecto.get("tir",0)
    tasa = proyecto.get("tasa",0)
    margen = tir - tasa if tir !=0 else 0
    payback = proyecto.get("payback",999)
    if isinstance(payback,str):
        try:
            payback = float(payback.split()[0])
        except:
            payback = 999
    vida = proyecto.get("años",1)
    if volatilidad < 0.2 and margen > 0.05 and payback < vida/2:
        return "Bajo"
    elif volatilidad > 0.5 or margen < 0:
        return "Alto"
    else:
        return "Medio"

def validaciones_inteligentes(proyecto):
    advertencias = []
    inv = proyecto.get("inversion",0)
    flujos = proyecto.get("flujos",[])
    suma_flujos = sum(flujos)
    if suma_flujos < inv:
        advertencias.append(f"La suma de flujos (${suma_flujos:,.2f}) es menor que la inversión (${inv:,.2f}).")
    if proyecto.get("tasa",0) > 0.5:
        advertencias.append("Tasa de descuento > 50%, muy sensible.")
    if any(f<0 for f in flujos):
        advertencias.append("Existen flujos de caja negativos.")
    return advertencias

def analisis_profundo_para_reporte(proyecto, otros_proyectos=None):
    nombre = proyecto["nombre"]
    metodo = proyecto.get("metodo","VPN")
    vpn = proyecto.get("vpn",0)
    tir = proyecto.get("tir",0)
    tasa = proyecto.get("tasa",0)
    riesgo = proyecto.get("riesgo","Medio")
    payback = proyecto.get("payback","N/A")
    inversion = proyecto.get("inversion",0)
    años = proyecto.get("años",1)
    cae = proyecto.get("cae",0)

    texto = f"Informe de evaluación del proyecto: {nombre}\n\n"
    texto += "1. Contexto y datos básicos\n"
    texto += f"   El proyecto requiere una inversión inicial de ${inversion:,.2f} y tiene una vida útil estimada de {años} años. "
    texto += f"La tasa mínima atractiva de retorno (TMAR) establecida es del {tasa*100:.2f}%. "
    texto += f"El valor residual al final del período es de ${proyecto['salvage']:,.2f}. "
    texto += "A continuación se analizan los principales indicadores financieros.\n\n"

    if metodo == "VPN":
        texto += "2. Valor Presente Neto (VPN)\n"
        texto += f"   El VPN calculado es de ${vpn:,.2f}. "
        if vpn > 0:
            texto += "Este valor es positivo, lo que significa que el proyecto genera valor económico por encima del costo de capital. "
            if vpn > inversion * 0.2:
                texto += "La rentabilidad es muy significativa, ya que el VPN supera el 20% de la inversión inicial. "
            else:
                texto += "La rentabilidad es moderada pero suficiente para justificar la inversión. "
        else:
            texto += "Este valor es negativo, lo que indica que el proyecto destruye valor y no alcanza a cubrir el costo de capital. "
        if otros_proyectos and len(otros_proyectos) >= 2:
            otros_vpns = [p["vpn"] for p in otros_proyectos if p["nombre"] != nombre]
            if otros_vpns:
                vpn_medio = np.mean(otros_vpns)
                if vpn > vpn_medio:
                    texto += f"En comparación con otros proyectos, este VPN es {((vpn - vpn_medio)/vpn_medio)*100:.1f}% superior al promedio. "
                else:
                    texto += f"Este VPN es inferior al promedio de otros proyectos en {((vpn_medio - vpn)/vpn_medio)*100:.1f}%. "
        texto += "\n\n"
        texto += "3. Tasa Interna de Retorno (TIR)\n"
        if tir < 0.001:
            texto += f"   La TIR obtenida es prácticamente 0%. Ello contrasta con el VPN {'positivo' if vpn > 0 else 'negativo'}. "
            if vpn > 0:
                texto += "A pesar del VPN positivo, una TIR cercana a cero indica que la rentabilidad efectiva es casi nula, lo que contradice una decisión de inversión. "
            else:
                texto += "Una TIR tan baja refuerza la decisión de rechazar el proyecto.\n"
        else:
            if tir > tasa:
                texto += f"   La TIR es {tir*100:.2f}%, superior a la TMAR ({tasa*100:.2f}%). Por tanto, el proyecto ofrece una rentabilidad anual mayor a la exigida. "
                if (tir - tasa) > 0.05:
                    texto += "El margen de seguridad es holgado ( >5 puntos), lo que protege ante cambios en la tasa de descuento.\n"
                else:
                    texto += "El margen sobre la TMAR es ajustado; un pequeño aumento en la tasa podría hacer el proyecto no rentable.\n"
            else:
                texto += f"   La TIR es {tir*100:.2f}%, inferior a la TMAR ({tasa*100:.2f}%). Por lo tanto, el proyecto no alcanza la rentabilidad mínima exigida.\n"
        texto += "\n"
        texto += "4. Costo Anual Equivalente (CAE)\n"
        texto += f"   El CAE del proyecto es de ${cae:,.2f} anuales. Este indicador permite comparar proyectos de diferente vida útil.\n\n"
    elif metodo == "CAE":
        texto += "2. Costo Anual Equivalente (CAE)\n"
        texto += f"   El CAE del proyecto es de ${cae:,.2f} anuales. "
        if cae < 0:
            texto += "El CAE es negativo, lo que indica que los costos anuales superan a los ingresos anuales equivalentes. El proyecto no es rentable.\n"
        else:
            texto += "El CAE es positivo, lo que indica ingresos netos anuales. Se debe comparar con otras alternativas.\n\n"
    else:  # TIR
        texto += "2. Tasa Interna de Retorno (TIR)\n"
        texto += f"   La TIR es {tir*100:.2f}%. "
        if tir > tasa:
            texto += f"Supera la TMAR ({tasa*100:.2f}%), por lo que el proyecto es rentable.\n"
            if (tir - tasa) > 0.05:
                texto += "   El margen de seguridad es holgado (mayor a 5 puntos porcentuales).\n"
        else:
            texto += f"Es inferior a la TMAR ({tasa*100:.2f}%), por lo que el proyecto no es rentable.\n\n"

    texto += "5. Periodo de recuperación y análisis de riesgo\n"
    payback_num = None
    if payback != "N/A":
        try:
            payback_num = float(payback.split()[0])
        except:
            pass
    if payback_num is not None:
        if payback_num < años/2:
            texto += f"   El payback simple es de {payback_num:.2f} años, menos de la mitad de la vida útil. La inversión se recupera rápidamente, reduciendo el riesgo. "
        elif payback_num <= años:
            texto += f"   El payback es de {payback_num:.2f} años, dentro de la vida útil pero no especialmente rápido. El riesgo asociado es moderado. "
        else:
            texto += f"   El payback supera la vida útil del proyecto, lo que implica que nunca se recupera totalmente la inversión durante la vida del proyecto. Riesgo alto. "
    else:
        texto += "   No se ha podido calcular un payback porque la suma de los flujos no alcanza la inversión inicial. Esto es indicio de baja rentabilidad o datos inconsistentes. "
    texto += f"La clasificación de riesgo global es {riesgo}. "
    if riesgo == "Bajo":
        texto += "Esto se debe a la baja volatilidad de los flujos, margen TIR-TMAR amplio y payback rápido. Proyecto conservador."
    elif riesgo == "Medio":
        texto += "Existe incertidumbre moderada, pero la rentabilidad esperada compensa parcialmente el riesgo."
    else:
        texto += "Alta volatilidad o margen de rentabilidad estrecho. Proyecto sensible a cambios externos. No recomendado para perfiles conservadores.\n"
    texto += "\n"
    texto += "6. Recomendación final\n"
    if metodo == "VPN":
        if vpn > 0 and tir > tasa and riesgo != "Alto":
            texto += "   El proyecto es viable y recomendable. Ofrece una rentabilidad atractiva con un nivel de riesgo manejable. Se aconseja su aceptación."
        elif vpn > 0 and riesgo == "Alto":
            texto += "   Aunque el VPN es positivo, el alto riesgo desaconseja la inversión. Se sugiere buscar alternativas con menor volatilidad."
        elif vpn > 0 and tir <= tasa:
            texto += "   El VPN es positivo pero la TIR no alcanza la TMAR. Esto puede deberse a la concentración de ingresos al final del período. No es un proyecto claramente aceptable; revise los datos."
        else:
            texto += "   El proyecto no cumple con los criterios básicos de rentabilidad. Se recomienda su rechazo."
    elif metodo == "CAE":
        if cae < 0:
            texto += "   Se recomienda rechazar el proyecto (costos netos anuales)."
        else:
            texto += "   El proyecto es aceptable según el CAE. Compárelo con otras alternativas para tomar la mejor decisión."
    else:
        if tir > tasa and riesgo != "Alto":
            texto += "   Se recomienda aceptar el proyecto. La rentabilidad supera la mínima exigida."
        elif tir > tasa and riesgo == "Alto":
            texto += "   Se recomienda aceptar con precaución debido al alto riesgo."
        else:
            texto += "   Se recomienda rechazar el proyecto. La rentabilidad no alcanza la tasa mínima."
    return texto

# ================== FUNCIONES PARA CALCULADORA UNIVERSAL ==================
def valor_futuro_simple(va, tasa, periodos):
    return va * (1 + tasa) ** periodos

def valor_presente_simple(vf, tasa, periodos):
    return vf / ((1 + tasa) ** periodos)

def tasa_implícita_simple(va, vf, periodos):
    return (vf / va) ** (1/periodos) - 1

def periodos_simple(va, vf, tasa):
    return math.log(vf / va) / math.log(1 + tasa)

def valor_futuro_anualidad(pago, tasa, periodos, anticipada=False):
    factor = ((1 + tasa) ** periodos - 1) / tasa
    if anticipada:
        factor *= (1 + tasa)
    return pago * factor

def valor_presente_anualidad(pago, tasa, periodos, anticipada=False):
    factor = (1 - (1 + tasa) ** -periodos) / tasa
    if anticipada:
        factor *= (1 + tasa)
    return pago * factor

def pago_anualidad_vf(vf, tasa, periodos, anticipada=False):
    factor = ((1 + tasa) ** periodos - 1) / tasa
    if anticipada:
        factor *= (1 + tasa)
    return vf / factor

def pago_anualidad_vp(vp, tasa, periodos, anticipada=False):
    factor = (1 - (1 + tasa) ** -periodos) / tasa
    if anticipada:
        factor *= (1 + tasa)
    return vp / factor

def periodos_anualidad_vf(vf, pago, tasa, anticipada=False):
    low, high = 1, 10000
    for _ in range(100):
        mid = (low+high)/2
        vf_calc = valor_futuro_anualidad(pago, tasa, mid, anticipada)
        if abs(vf_calc - vf) < 0.01:
            return mid
        if vf_calc < vf:
            low = mid
        else:
            high = mid
    return (low+high)/2

def periodos_anualidad_vp(vp, pago, tasa, anticipada=False):
    low, high = 1, 10000
    for _ in range(100):
        mid = (low+high)/2
        vp_calc = valor_presente_anualidad(pago, tasa, mid, anticipada)
        if abs(vp_calc - vp) < 0.01:
            return mid
        if vp_calc < vp:
            low = mid
        else:
            high = mid
    return (low+high)/2

def tasa_anualidad_vp(vp, pago, periodos, anticipada=False):
    low, high = 0.0, 1.0
    for _ in range(100):
        mid = (low+high)/2
        vp_calc = valor_presente_anualidad(pago, mid, periodos, anticipada)
        if abs(vp_calc - vp) < 0.1:
            return mid
        if vp_calc > vp:
            low = mid
        else:
            high = mid
    return (low+high)/2

def tasa_anualidad_vf(vf, pago, periodos, anticipada=False):
    low, high = 0.0, 1.0
    for _ in range(100):
        mid = (low+high)/2
        vf_calc = valor_futuro_anualidad(pago, mid, periodos, anticipada)
        if abs(vf_calc - vf) < 0.1:
            return mid
        if vf_calc < vf:
            low = mid
        else:
            high = mid
    return (low+high)/2

def conversion_tasa_nominal_a_efectiva(nominal, m):
    return (1 + nominal / m) ** m - 1

def conversion_tasa_efectiva_a_nominal(efectiva, m):
    return m * ((1 + efectiva) ** (1 / m) - 1)

# ================== BASE DE DATOS ==================
DB_NAME = "seae_final.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS proyectos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT,
        metodo TEXT,
        inversion REAL,
        flujos TEXT,
        años INTEGER,
        tasa REAL,
        salvage REAL,
        vpn REAL,
        tir REAL,
        cae REAL,
        payback TEXT,
        riesgo TEXT,
        recomendacion TEXT,
        fecha_creacion TEXT
    )''')
    conn.commit()
    conn.close()

def guardar_proyecto(proyecto):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    flujos_json = json.dumps(proyecto["flujos"])
    c.execute('''INSERT INTO proyectos 
                 (nombre, metodo, inversion, flujos, años, tasa, salvage, vpn, tir, cae, payback, riesgo, recomendacion, fecha_creacion)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
              (proyecto["nombre"], proyecto.get("metodo","VPN"), proyecto["inversion"], flujos_json, proyecto["años"],
               proyecto["tasa"], proyecto["salvage"], proyecto.get("vpn",0), proyecto.get("tir",0),
               proyecto.get("cae",0), proyecto.get("payback","N/A"), proyecto.get("riesgo","Medio"),
               proyecto.get("recomendacion",""), datetime.now().isoformat()))
    conn.commit()
    conn.close()

def cargar_todos_proyectos():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id, nombre, metodo, inversion, flujos, años, tasa, salvage, vpn, tir, cae, payback, riesgo, recomendacion, fecha_creacion FROM proyectos ORDER BY fecha_creacion DESC")
    rows = c.fetchall()
    conn.close()
    proyectos = []
    for row in rows:
        proyectos.append({
            "id": row[0], "nombre": row[1], "metodo": row[2], "inversion": row[3],
            "flujos": json.loads(row[4]), "años": row[5], "tasa": row[6],
            "salvage": row[7], "vpn": row[8], "tir": row[9], "cae": row[10],
            "payback": row[11], "riesgo": row[12], "recomendacion": row[13], "fecha": row[14]
        })
    return proyectos

def eliminar_proyecto(id_proy):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM proyectos WHERE id = ?", (id_proy,))
    conn.commit()
    conn.close()

def obtener_proyecto_por_id(id_proy):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id, nombre, metodo, inversion, flujos, años, tasa, salvage, vpn, tir, cae, payback, riesgo, recomendacion FROM proyectos WHERE id=?", (id_proy,))
    row = c.fetchone()
    conn.close()
    if row:
        return {
            "id": row[0], "nombre": row[1], "metodo": row[2], "inversion": row[3],
            "flujos": json.loads(row[4]), "años": row[5], "tasa": row[6],
            "salvage": row[7], "vpn": row[8], "tir": row[9], "cae": row[10],
            "payback": row[11], "riesgo": row[12], "recomendacion": row[13]
        }
    return None

# ================== gráficos según método ==================
def generar_graficos_reporte(proyecto):
    imagenes = []
    metodo = proyecto.get("metodo", "VPN")
    inv = proyecto["inversion"]
    tasa = proyecto["tasa"]
    años = proyecto["años"]
    salvage = proyecto["salvage"]
    flujos = proyecto["flujos"][:años]
    vpn, flujos_desc = calcular_vpn(inv, flujos, tasa, años, salvage)

    if metodo == "VPN":
        fig1, ax1 = plt.subplots(figsize=(4, 2.5), dpi=80)
        ax1.bar([proyecto["nombre"]], [vpn], color=tema.success if vpn > 0 else tema.danger)
        ax1.axhline(y=0, color='black', linestyle='--')
        ax1.set_title("VPN del Proyecto")
        ax1.set_ylabel("VPN ($)")
        fig1.tight_layout()
        fig1.savefig("temp_vpn.png")
        plt.close(fig1)
        imagenes.append("temp_vpn.png")
        if flujos_desc:
            fig2, ax2 = plt.subplots(figsize=(4, 2.5), dpi=80)
            acum = np.cumsum(flujos_desc)
            ax2.plot(range(1, len(acum)+1), acum, marker='o', color=tema.accent)
            ax2.axhline(y=0, color='red', linestyle='--')
            ax2.fill_between(range(1, len(acum)+1), 0, acum,
                             where=(acum > 0), color='green', alpha=0.3)
            ax2.fill_between(range(1, len(acum)+1), 0, acum,
                             where=(acum < 0), color='red', alpha=0.3)
            ax2.set_title("Evolución del VPN Acumulado")
            ax2.set_xlabel("Año")
            ax2.set_ylabel("VPN Acumulado ($)")
            fig2.tight_layout()
            fig2.savefig("temp_acum.png")
            plt.close(fig2)
            imagenes.append("temp_acum.png")
    elif metodo == "CAE":
        cae = calcular_cae(vpn, tasa, años)
        fig, ax = plt.subplots(figsize=(5, 3), dpi=80)
        ax.bar(range(1, años+1), flujos, color=tema.accent, label="Flujo anual neto")
        ax.axhline(y=cae, color='red', linestyle='--', linewidth=2,
                   label=f"CAE = ${cae:,.2f}")
        ax.set_title("Flujos Anuales y Costo Anual Equivalente (CAE)")
        ax.set_xlabel("Año")
        ax.set_ylabel("Monto ($)")
        ax.legend()
        fig.tight_layout()
        fig.savefig("temp_cae.png")
        plt.close(fig)
        imagenes.append("temp_cae.png")
    elif metodo == "TIR":
        tir = calcular_tir(inv, flujos, años, salvage)
        tasas = np.linspace(0.001, max(tasa*2, 0.3), 50)
        vpns = [calcular_vpn(inv, flujos, t, años, salvage)[0] for t in tasas]
        fig, ax = plt.subplots(figsize=(5, 3), dpi=80)
        ax.plot(tasas, vpns, color=tema.accent, linewidth=2)
        if tir is not None:
            ax.axvline(x=tir, color='red', linestyle='--',
                       label=f"TIR = {tir*100:.2f}%")
        ax.axhline(y=0, color='black', linestyle='--')
        ax.set_xlabel("Tasa de descuento")
        ax.set_ylabel("VPN ($)")
        ax.set_title("Perfil del VPN — Determinación de TIR")
        ax.legend()
        fig.tight_layout()
        fig.savefig("temp_tir.png")
        plt.close(fig)
        imagenes.append("temp_tir.png")
    return imagenes

# ================== REPORTES ==================
class PDFReport(FPDF):
    def header(self):
        if self.page_no() == 1:
            self.set_font('Arial', 'B', 16)
            self.set_text_color(0,51,102)
            self.cell(0,10,'UNIVERSIDAD DE EL SALVADOR',0,1,'C')
            self.set_font('Arial','',12)
            self.set_text_color(0,0,0)
            self.cell(0,6,'Facultad Multidisciplinaria De Occidente',0,1,'C')
            self.cell(0,6,'Departamento de Ingeniería y Arquitectura',0,1,'C')
            self.ln(8)
            self.set_font('Arial','B',14)
            self.set_text_color(0,51,102)
            self.cell(0,10,'INFORME DE EVALUACIÓN ECONÓMICA',0,1,'C')
            self.line(10,self.get_y(),200,self.get_y())
            self.ln(5)
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial','I',8)
        self.set_text_color(128)
        self.cell(0,10,f'Página {self.page_no()} - SEAE',0,0,'C')

def generar_reporte_pdf_avanzado(proyecto, otros_proyectos, ruta):
    imagenes = generar_graficos_reporte(proyecto)
    pdf = PDFReport()
    pdf.add_page()
    pdf.set_font('Arial','B',11)
    pdf.cell(0,8,'DATOS DEL PROYECTO',0,1)
    pdf.set_font('Arial','',10)
    pdf.cell(0,6,f"   Nombre: {proyecto['nombre']}",0,1)
    pdf.cell(0,6,f"   Método: {proyecto.get('metodo','VPN')}",0,1)
    pdf.cell(0,6,f"   Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}",0,1)
    pdf.cell(0,6,f"   Inversión: ${proyecto['inversion']:,.2f}",0,1)
    pdf.cell(0,6,f"   Vida útil: {proyecto['años']} años",0,1)
    pdf.cell(0,6,f"   TMAR: {proyecto['tasa']*100:.2f}%",0,1)
    pdf.cell(0,6,f"   Valor residual: ${proyecto['salvage']:,.2f}",0,1)
    pdf.ln(4)
    for img in imagenes:
        if os.path.exists(img):
            pdf.image(img, x=30, y=pdf.get_y(), w=130)
            pdf.ln(45)
    pdf.add_page()
    pdf.set_font('Arial','B',11)
    pdf.cell(0,8,'CONCLUSIÓN Y ANÁLISIS',0,1)
    pdf.set_font('Arial','',10)
    analisis = analisis_profundo_para_reporte(proyecto, otros_proyectos)
    pdf.multi_cell(0,6,analisis)
    pdf.output(ruta)
    for img in imagenes:
        if os.path.exists(img):
            os.remove(img)

def generar_reporte_excel(proyecto, ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws['A1'] = "SEAE - Evaluación Económica"
    ws['A1'].font = Font(bold=True,size=14)
    ws['A3'] = "Proyecto:"; ws['B3'] = proyecto['nombre']
    ws['A4'] = "Método:"; ws['B4'] = proyecto.get('metodo','VPN')
    ws['A5'] = "Inversión inicial ($):"; ws['B5'] = proyecto['inversion']
    ws['A6'] = "Vida útil (años):"; ws['B6'] = proyecto['años']
    ws['A7'] = "TMAR (%):"; ws['B7'] = proyecto['tasa']*100
    ws['A8'] = "Valor residual ($):"; ws['B8'] = proyecto['salvage']
    ws['A10'] = "INDICADORES"; ws['A10'].font = Font(bold=True)
    ws['A11'] = "VPN ($):"; ws['B11'] = proyecto['vpn']
    ws['A12'] = "TIR (%):"; ws['B12'] = proyecto['tir']*100
    ws['A13'] = "CAE ($):"; ws['B13'] = proyecto['cae']
    ws['A14'] = "Payback:"; ws['B14'] = proyecto['payback']
    ws['A15'] = "Riesgo:"; ws['B15'] = proyecto['riesgo']
    ws['A17'] = "Recomendación:"; ws['B17'] = proyecto['recomendacion']
    wb.save(ruta)

def generar_reporte_comparativo_pdf(proyectos, ruta):
    if not proyectos: return
    pdf = PDFReport()
    pdf.add_page()
    pdf.set_font('Arial','B',14)
    pdf.cell(0,10,'INFORME COMPARATIVO GLOBAL',0,1,'C')
    pdf.ln(5)
    pdf.set_font('Arial','B',10)
    pdf.cell(40,8,'Proyecto',1,0,'C')
    pdf.cell(30,8,'Método',1,0,'C')
    pdf.cell(50,8,'Resultado',1,0,'C')
    pdf.cell(40,8,'Riesgo',1,1,'C')
    for p in proyectos:
        metodo = p.get('metodo','VPN')
        if metodo == "VPN":
            res = f"${p['vpn']:,.2f}"
        elif metodo == "CAE":
            res = f"${p['cae']:,.2f} anual"
        else:
            res = f"{p['tir']*100:.2f}%"
        pdf.cell(40,8,p['nombre'],1,0)
        pdf.cell(30,8,metodo,1,0)
        pdf.cell(50,8,res,1,0)
        pdf.cell(40,8,p['riesgo'],1,1,'C')
    pdf.output(ruta)

def generar_reporte_calculadora(datos, titulo, analisis, grafico, ruta):
    pdf = PDFReport()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 8, titulo, 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 6, f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1)
    pdf.ln(4)
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(0, 8, 'DATOS DE ENTRADA', 0, 1)
    pdf.set_font('Arial', '', 10)
    for k, v in datos.items():
        if v is not None:
            pdf.cell(0, 6, f"   {k}: {v}", 0, 1)
    pdf.ln(4)
    if grafico and os.path.exists(grafico):
        pdf.image(grafico, x=30, y=pdf.get_y(), w=130)
        pdf.ln(50)
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(0, 8, 'ANÁLISIS DEL RESULTADO', 0, 1)
    pdf.set_font('Arial', '', 10)
    pdf.multi_cell(0, 6, analisis)
    pdf.output(ruta)
    if grafico and os.path.exists(grafico):
        os.remove(grafico)

# ================== FUNCIÓN DE DECISIÓN SEGÚN MÉTODO ==================
def decidir_aceptacion(proyecto):
    metodo = proyecto.get("metodo", "VPN")
    if metodo == "VPN":
        return proyecto["vpn"] > 0
    elif metodo == "TIR":
        return proyecto["tir"] > proyecto["tasa"]
    else:  # CAE
        return proyecto["cae"] > 0

# ================== TOOLTIP ==================
class ToolTip:
    def __init__(self,widget,text):
        self.widget=widget
        self.text=text
        self.tip_window=None
        widget.bind('<Enter>',self._enter)
        widget.bind('<Leave>',self._leave)
    def _enter(self,event):
        x,y,_,_ = self.widget.bbox("insert")
        x+=self.widget.winfo_rootx()+25
        y+=self.widget.winfo_rooty()+20
        self.tip_window = tw = ctk.CTkToplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.geometry(f"+{x}+{y}")
        label = ctk.CTkLabel(tw,text=self.text,justify="left",corner_radius=5,fg_color="#333",text_color="white")
        label.pack()
    def _leave(self,event):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window=None

# ================== CLASE PRINCIPAL ==================
class SEAE_Completo(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("SEAE - SISTEMA DE EVALUACIÓN DE ALTERNATIVAS ECONÓMICAS ")
        self.geometry("1400x900")
        self.minsize(1200,800)
        self.configure(bg=tema.bg)
        self.protocol("WM_DELETE_WINDOW",self.on_closing)

        self.proyectos = []
        self.project_frames = []

        self.top_frame = ctk.CTkFrame(self,height=60,corner_radius=0,fg_color=tema.card)
        self.top_frame.pack(fill='x',side='top')
        self.top_frame.pack_propagate(False)
        self.title_label = ctk.CTkLabel(self.top_frame,text="SEAE - SISTEMA DE EVALUACIÓN DE ALTERNATIVAS ECONÓMICAS ",
                                        font=("Segoe UI",18,"bold"),text_color=tema.text_primary)
        self.title_label.pack(side='left',padx=30,pady=15)

        self.tabview = ctk.CTkTabview(self,corner_radius=15,
                                      segmented_button_fg_color=tema.border,
                                      segmented_button_selected_color=tema.accent,
                                      segmented_button_unselected_color=tema.card)
        self.tabview.pack(fill='both',expand=True,padx=20,pady=20)
        tabs = ["Calculadora","Proyectos","Dashboard","Gráficos","Comparación","Sensibilidad","Historial","Glosario","Ayuda"]
        for t in tabs:
            self.tabview.add(t)

        init_db()
        self.cargar_proyectos_guardados()

        self.build_calculadora_tab()
        self.build_proyectos_tab()
        self.build_dashboard_tab()
        self.build_graficos_tab()
        self.build_comparacion_tab()
        self.build_sensibilidad_tab()
        self.build_historial_tab()
        self.build_glosario_tab()
        self.build_ayuda_tab()

        self.status_bar = ctk.CTkLabel(self,text="Sistema listo",anchor="w",
                                       fg_color=tema.border,text_color=tema.text_secondary,height=30)
        self.status_bar.pack(fill='x',side='bottom')
        self.after(100,lambda: self.tabview.set("Calculadora"))

    def on_closing(self):
        self.destroy()

    def cargar_proyectos_guardados(self):
        self.proyectos = cargar_todos_proyectos()
        for p in self.proyectos:
            p["res_label"]=None; p["warn_label"]=None
            p["flujo_entries"]=[]; p["flujos_container"]=None
            p["after_id"] = None

    def guardar_y_actualizar(self):
        self.actualizar_dashboard()
        self.actualizar_graficos_tab()
        self.actualizar_comparacion_tab()
        self.cargar_historial_tabla()
        self.cargar_proyectos_en_ui()

    # ================== PESTAÑA CALCULADORA ==================
    def build_calculadora_tab(self):
        tab = self.tabview.tab("Calculadora")
        sub_tab = ctk.CTkTabview(tab,corner_radius=10,
                                 segmented_button_fg_color=tema.border,
                                 segmented_button_selected_color=tema.accent,
                                 segmented_button_unselected_color=tema.card)
        sub_tab.pack(fill='both',expand=True,padx=10,pady=10)
        sub_tab.add("Valor del Dinero")
        sub_tab.add("Anualidades")
        sub_tab.add("Amortización")
        sub_tab.add("Conversión de Tasas")
        sub_tab.add("VPN / CAE / TIR")
        self.build_sub_valor_dinero(sub_tab.tab("Valor del Dinero"))
        self.build_sub_anualidades(sub_tab.tab("Anualidades"))
        self.build_sub_amortizacion(sub_tab.tab("Amortización"))
        self.build_sub_conversion_tasas(sub_tab.tab("Conversión de Tasas"))
        self.build_sub_vpn_cae_tir(sub_tab.tab("VPN / CAE / TIR"))

    # ---------- Valor del Dinero ----------
    def build_sub_valor_dinero(self,parent):
        frame = ctk.CTkFrame(parent,fg_color=tema.card,corner_radius=10)
        frame.pack(fill='both',expand=True,padx=10,pady=10)
        ctk.CTkLabel(frame,text="Valor Actual (VA):").grid(row=0,column=0,padx=10,pady=5,sticky="e")
        self.vd_va = ctk.CTkEntry(frame,width=150); self.vd_va.grid(row=0,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Valor Futuro (VF):").grid(row=1,column=0,padx=10,pady=5,sticky="e")
        self.vd_vf = ctk.CTkEntry(frame,width=150); self.vd_vf.grid(row=1,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Tasa por período (%):").grid(row=2,column=0,padx=10,pady=5,sticky="e")
        self.vd_tasa = ctk.CTkEntry(frame,width=150); self.vd_tasa.grid(row=2,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Número de períodos:").grid(row=3,column=0,padx=10,pady=5,sticky="e")
        self.vd_n = ctk.CTkEntry(frame,width=150); self.vd_n.grid(row=3,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="(Dejar vacío el campo a calcular)",font=("Segoe UI",10)).grid(row=4,column=0,columnspan=2,pady=5)
        btn_calc = ctk.CTkButton(frame,text="Calcular",command=self.calcular_valor_dinero,fg_color=tema.success)
        btn_calc.grid(row=5,column=0,columnspan=2,pady=10)
        self.vd_resultado = ctk.CTkTextbox(frame,height=120,wrap="word")
        self.vd_resultado.grid(row=6,column=0,columnspan=2,padx=10,pady=10,sticky="ew")
        btn_reporte = ctk.CTkButton(frame,text="Generar reporte PDF",command=self.reporte_valor_dinero,fg_color=tema.accent)
        btn_reporte.grid(row=7,column=0,columnspan=2,pady=5)

    def calcular_valor_dinero(self):
        try:
            va = self.vd_va.get().strip()
            vf = self.vd_vf.get().strip()
            tasa = self.vd_tasa.get().strip()
            n = self.vd_n.get().strip()
            va_val = float(va) if va else None
            vf_val = float(vf) if vf else None
            tasa_val = float(tasa)/100 if tasa else None
            n_val = int(n) if n else None
            if va_val is None:
                if vf_val is None or tasa_val is None or n_val is None:
                    self.vd_resultado.delete("0.0","end")
                    self.vd_resultado.insert("0.0","Faltan datos: debe ingresar VF, tasa y n para calcular VA.")
                    return
                res = valor_presente_simple(vf_val, tasa_val, n_val)
                analisis = f"Se calculó el valor presente (VA) a partir del valor futuro de ${vf_val:,.2f}, una tasa del {tasa_val*100:.2f}% y {n_val} períodos. La fórmula utilizada es VA = VF / (1 + i)^n. Sustituyendo: VA = {vf_val:,.2f} / (1 + {tasa_val:.4f})^{n_val} = {res:,.2f}. Este es el monto que debería invertirse hoy para alcanzar el valor futuro deseado."
                self.vd_ultimo = ({"Valor Actual": None, "Valor Futuro": vf_val, "Tasa %": tasa, "Períodos": n_val}, f"VA = {res:,.2f}", analisis)
                self.vd_resultado.delete("0.0","end")
                self.vd_resultado.insert("0.0",f"Valor Actual = {res:,.2f}")
            elif vf_val is None:
                if va_val is None or tasa_val is None or n_val is None:
                    self.vd_resultado.delete("0.0","end")
                    self.vd_resultado.insert("0.0","Faltan datos: debe ingresar VA, tasa y n para calcular VF.")
                    return
                res = valor_futuro_simple(va_val, tasa_val, n_val)
                analisis = f"Se calculó el valor futuro (VF) a partir de un valor actual de ${va_val:,.2f}, una tasa del {tasa_val*100:.2f}% y {n_val} períodos. La fórmula es VF = VA * (1 + i)^n. Sustituyendo: VF = {va_val:,.2f} * (1 + {tasa_val:.4f})^{n_val} = {res:,.2f}. Este es el monto acumulado al final de los {n_val} períodos."
                self.vd_ultimo = ({"Valor Actual": va_val, "Valor Futuro": None, "Tasa %": tasa, "Períodos": n_val}, f"VF = {res:,.2f}", analisis)
                self.vd_resultado.delete("0.0","end")
                self.vd_resultado.insert("0.0",f"Valor Futuro = {res:,.2f}")
            elif tasa_val is None:
                if va_val is None or vf_val is None or n_val is None:
                    self.vd_resultado.delete("0.0","end")
                    self.vd_resultado.insert("0.0","Faltan datos: debe ingresar VA, VF y n para calcular tasa.")
                    return
                res = tasa_implícita_simple(va_val, vf_val, n_val)*100
                analisis = f"Se calculó la tasa de interés implícita a partir de un valor actual de ${va_val:,.2f}, un valor futuro de ${vf_val:,.2f} y {n_val} períodos. La fórmula es i = (VF/VA)^(1/n) - 1. Sustituyendo: i = ({vf_val:,.2f}/{va_val:,.2f})^(1/{n_val}) - 1 = {res:.4f}%. Esta es la tasa efectiva por período que iguala ambos valores."
                self.vd_ultimo = ({"Valor Actual": va_val, "Valor Futuro": vf_val, "Tasa %": None, "Períodos": n_val}, f"Tasa = {res:.4f}%", analisis)
                self.vd_resultado.delete("0.0","end")
                self.vd_resultado.insert("0.0",f"Tasa por período = {res:.4f}%")
            elif n_val is None:
                if va_val is None or vf_val is None or tasa_val is None:
                    self.vd_resultado.delete("0.0","end")
                    self.vd_resultado.insert("0.0","Faltan datos: debe ingresar VA, VF y tasa para calcular n.")
                    return
                res = periodos_simple(va_val, vf_val, tasa_val)
                analisis = f"Se calculó el número de períodos necesarios para que ${va_val:,.2f} se conviertan en ${vf_val:,.2f} a una tasa del {tasa_val*100:.2f}%. La fórmula es n = ln(VF/VA) / ln(1+i). Sustituyendo: n = ln({vf_val/va_val:.4f}) / ln(1+{tasa_val:.4f}) = {res:.4f}."
                self.vd_ultimo = ({"Valor Actual": va_val, "Valor Futuro": vf_val, "Tasa %": tasa, "Períodos": None}, f"n = {res:.4f}", analisis)
                self.vd_resultado.delete("0.0","end")
                self.vd_resultado.insert("0.0",f"Número de períodos = {res:.4f}")
            else:
                self.vd_resultado.delete("0.0","end")
                self.vd_resultado.insert("0.0","Todos los campos llenos. Deje vacío el que desea calcular.")
        except Exception as e:
            self.vd_resultado.delete("0.0","end")
            self.vd_resultado.insert("0.0",f"Error: {str(e)}")

    def reporte_valor_dinero(self):
        if not hasattr(self, 'vd_ultimo'):
            messagebox.showwarning("Sin datos", "Primero realice un cálculo.")
            return
        datos, res, analisis = self.vd_ultimo
        if datos.get("Valor Actual") is not None and datos.get("Tasa %") is not None and datos.get("Períodos") is not None:
            va = datos["Valor Actual"]
            tasa = float(datos["Tasa %"])/100
            n = datos["Períodos"]
            años = list(range(n+1))
            montos = [va * (1+tasa)**t for t in años]
            fig, ax = plt.subplots(figsize=(5,3), dpi=90)
            ax.plot(años, montos, marker='o', color=tema.accent)
            ax.fill_between(años, montos, alpha=0.2, color=tema.accent)
            ax.set_title("Crecimiento del capital")
            ax.set_xlabel("Período")
            ax.set_ylabel("Monto ($)")
            fig.tight_layout()
            grafico = "temp_vd.png"
            fig.savefig(grafico)
            plt.close(fig)
        else:
            grafico = None
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if archivo:
            generar_reporte_calculadora(datos, "CÁLCULO DE VALOR DEL DINERO", analisis, grafico, archivo)
            messagebox.showinfo("Reporte", f"PDF guardado en {archivo}")

    # ---------- Anualidades ----------
    def build_sub_anualidades(self,parent):
        frame = ctk.CTkFrame(parent,fg_color=tema.card,corner_radius=10)
        frame.pack(fill='both',expand=True,padx=10,pady=10)
        ctk.CTkLabel(frame,text="Tipo:").grid(row=0,column=0,padx=10,pady=5,sticky="e")
        self.an_tipo = ctk.CTkOptionMenu(frame,values=["Ordinaria","Anticipada"],width=120)
        self.an_tipo.grid(row=0,column=1,padx=10,pady=5,sticky="w")
        ctk.CTkLabel(frame,text="Valor Presente (VP):").grid(row=1,column=0,padx=10,pady=5,sticky="e")
        self.an_vp = ctk.CTkEntry(frame,width=150); self.an_vp.grid(row=1,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Valor Futuro (VF):").grid(row=2,column=0,padx=10,pady=5,sticky="e")
        self.an_vf = ctk.CTkEntry(frame,width=150); self.an_vf.grid(row=2,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Pago (PMT):").grid(row=3,column=0,padx=10,pady=5,sticky="e")
        self.an_pmt = ctk.CTkEntry(frame,width=150); self.an_pmt.grid(row=3,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Tasa por período (%):").grid(row=4,column=0,padx=10,pady=5,sticky="e")
        self.an_tasa = ctk.CTkEntry(frame,width=150); self.an_tasa.grid(row=4,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Número de períodos:").grid(row=5,column=0,padx=10,pady=5,sticky="e")
        self.an_n = ctk.CTkEntry(frame,width=150); self.an_n.grid(row=5,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="(Dejar vacío el campo a calcular)",font=("Segoe UI",10)).grid(row=6,column=0,columnspan=2,pady=5)
        btn_calc = ctk.CTkButton(frame,text="Calcular",command=self.calcular_anualidad,fg_color=tema.success)
        btn_calc.grid(row=7,column=0,columnspan=2,pady=10)
        self.an_resultado = ctk.CTkTextbox(frame,height=150,wrap="word")
        self.an_resultado.grid(row=8,column=0,columnspan=2,padx=10,pady=10,sticky="ew")
        btn_reporte = ctk.CTkButton(frame,text="Generar reporte PDF",command=self.reporte_anualidad,fg_color=tema.accent)
        btn_reporte.grid(row=9,column=0,columnspan=2,pady=5)

    def calcular_anualidad(self):
        try:
            anticipada = (self.an_tipo.get() == "Anticipada")
            vp = self.an_vp.get().strip()
            vf = self.an_vf.get().strip()
            pmt = self.an_pmt.get().strip()
            tasa = self.an_tasa.get().strip()
            n = self.an_n.get().strip()
            vp_val = float(vp) if vp else None
            vf_val = float(vf) if vf else None
            pmt_val = float(pmt) if pmt else None
            tasa_val = float(tasa)/100 if tasa else None
            n_val = int(n) if n else None
            if pmt_val is not None and tasa_val is not None and n_val is not None:
                if vp_val is None and vf_val is None:
                    res = valor_presente_anualidad(pmt_val, tasa_val, n_val, anticipada)
                    analisis = f"Se calculó el valor presente de una anualidad {self.an_tipo.get().lower()} con pagos de ${pmt_val:,.2f}, tasa {tasa_val*100:.2f}% y {n_val} períodos. El resultado es ${res:,.2f}. Este es el monto que, invertido hoy, generaría los pagos periódicos."
                    self.an_ultimo = ({"VP":None, "VF":None, "PMT":pmt_val, "Tasa %": tasa, "Períodos": n_val, "Tipo":self.an_tipo.get()}, f"VP = {res:,.2f}", analisis)
                    self.an_resultado.delete("0.0","end")
                    self.an_resultado.insert("0.0",f"Valor Presente = {res:,.2f}")
                elif vf_val is None:
                    res = valor_futuro_anualidad(pmt_val, tasa_val, n_val, anticipada)
                    analisis = f"Se calculó el valor futuro de una anualidad {self.an_tipo.get().lower()} con pagos de ${pmt_val:,.2f}, tasa {tasa_val*100:.2f}% y {n_val} períodos. El resultado es ${res:,.2f}. Este es el monto acumulado al final de los {n_val} períodos."
                    self.an_ultimo = ({"VP":None, "VF":None, "PMT":pmt_val, "Tasa %": tasa, "Períodos": n_val, "Tipo":self.an_tipo.get()}, f"VF = {res:,.2f}", analisis)
                    self.an_resultado.delete("0.0","end")
                    self.an_resultado.insert("0.0",f"Valor Futuro = {res:,.2f}")
                else:
                    self.an_resultado.delete("0.0","end")
                    self.an_resultado.insert("0.0","Datos completos. Deje vacío el valor a calcular.")
            elif vp_val is not None and tasa_val is not None and n_val is not None:
                res = pago_anualidad_vp(vp_val, tasa_val, n_val, anticipada)
                analisis = f"Se calculó el pago periódico de una anualidad {self.an_tipo.get().lower()} que tiene un valor presente de ${vp_val:,.2f}, una tasa del {tasa_val*100:.2f}% y {n_val} períodos. El pago resultante es ${res:,.2f}."
                self.an_ultimo = ({"VP":vp_val, "VF":None, "PMT":None, "Tasa %": tasa, "Períodos": n_val, "Tipo":self.an_tipo.get()}, f"PMT = {res:,.2f}", analisis)
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0",f"Pago = {res:,.2f}")
            elif vf_val is not None and tasa_val is not None and n_val is not None:
                res = pago_anualidad_vf(vf_val, tasa_val, n_val, anticipada)
                analisis = f"Se calculó el pago periódico de una anualidad {self.an_tipo.get().lower()} que tiene un valor futuro de ${vf_val:,.2f}, una tasa del {tasa_val*100:.2f}% y {n_val} períodos. El pago resultante es ${res:,.2f}."
                self.an_ultimo = ({"VP":None, "VF":vf_val, "PMT":None, "Tasa %": tasa, "Períodos": n_val, "Tipo":self.an_tipo.get()}, f"PMT = {res:,.2f}", analisis)
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0",f"Pago = {res:,.2f}")
            elif vp_val is not None and pmt_val is not None and n_val is not None:
                res = tasa_anualidad_vp(vp_val, pmt_val, n_val, anticipada)*100
                analisis = f"Se calculó la tasa de interés de una anualidad {self.an_tipo.get().lower()} con valor presente de ${vp_val:,.2f}, pago de ${pmt_val:,.2f} y {n_val} períodos. La tasa obtenida es {res:.4f}% por período."
                self.an_ultimo = ({"VP":vp_val, "VF":None, "PMT":pmt_val, "Tasa %": None, "Períodos": n_val, "Tipo":self.an_tipo.get()}, f"Tasa = {res:.4f}%", analisis)
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0",f"Tasa = {res:.4f}%")
            elif vf_val is not None and pmt_val is not None and n_val is not None:
                res = tasa_anualidad_vf(vf_val, pmt_val, n_val, anticipada)*100
                analisis = f"Se calculó la tasa de interés de una anualidad {self.an_tipo.get().lower()} con valor futuro de ${vf_val:,.2f}, pago de ${pmt_val:,.2f} y {n_val} períodos. La tasa obtenida es {res:.4f}% por período."
                self.an_ultimo = ({"VP":None, "VF":vf_val, "PMT":pmt_val, "Tasa %": None, "Períodos": n_val, "Tipo":self.an_tipo.get()}, f"Tasa = {res:.4f}%", analisis)
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0",f"Tasa = {res:.4f}%")
            elif vp_val is not None and pmt_val is not None and tasa_val is not None:
                res = periodos_anualidad_vp(vp_val, pmt_val, tasa_val, anticipada)
                analisis = f"Se calculó el número de períodos de una anualidad {self.an_tipo.get().lower()} con valor presente de ${vp_val:,.2f}, pago de ${pmt_val:,.2f} y tasa del {tasa_val*100:.2f}%. El resultado es {res:.4f} períodos."
                self.an_ultimo = ({"VP":vp_val, "VF":None, "PMT":pmt_val, "Tasa %": tasa, "Períodos": None, "Tipo":self.an_tipo.get()}, f"n = {res:.4f}", analisis)
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0",f"Períodos = {res:.4f}")
            elif vf_val is not None and pmt_val is not None and tasa_val is not None:
                res = periodos_anualidad_vf(vf_val, pmt_val, tasa_val, anticipada)
                analisis = f"Se calculó el número de períodos de una anualidad {self.an_tipo.get().lower()} con valor futuro de ${vf_val:,.2f}, pago de ${pmt_val:,.2f} y tasa del {tasa_val*100:.2f}%. El resultado es {res:.4f} períodos."
                self.an_ultimo = ({"VP":None, "VF":vf_val, "PMT":pmt_val, "Tasa %": tasa, "Períodos": None, "Tipo":self.an_tipo.get()}, f"n = {res:.4f}", analisis)
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0",f"Períodos = {res:.4f}")
            else:
                self.an_resultado.delete("0.0","end")
                self.an_resultado.insert("0.0","Datos insuficientes.")
        except Exception as e:
            self.an_resultado.delete("0.0","end")
            self.an_resultado.insert("0.0",f"Error: {str(e)}")

    def reporte_anualidad(self):
        if not hasattr(self, 'an_ultimo'):
            messagebox.showwarning("Sin datos", "Primero realice un cálculo.")
            return
        datos, res, analisis = self.an_ultimo
        if datos.get("PMT") is not None and datos.get("Períodos") is not None:
            pmt = datos["PMT"]
            n = datos["Períodos"]
            flujos = [pmt] * n
            fig, ax = plt.subplots(figsize=(5,3), dpi=90)
            ax.bar(range(1,n+1), flujos, color=tema.accent)
            ax.set_title("Flujos de la anualidad")
            ax.set_xlabel("Período")
            ax.set_ylabel("Monto ($)")
            fig.tight_layout()
            grafico = "temp_an.png"
            fig.savefig(grafico)
            plt.close(fig)
        else:
            grafico = None
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if archivo:
            generar_reporte_calculadora(datos, "CÁLCULO DE ANUALIDADES", analisis, grafico, archivo)
            messagebox.showinfo("Reporte", f"PDF guardado en {archivo}")

    # ---------- Amortización ----------
    def build_sub_amortizacion(self,parent):
        frame = ctk.CTkFrame(parent,fg_color=tema.card,corner_radius=10)
        frame.pack(fill='both',expand=True,padx=10,pady=10)
        ctk.CTkLabel(frame,text="Monto del préstamo:").grid(row=0,column=0,padx=10,pady=5,sticky="e")
        self.amo_monto = ctk.CTkEntry(frame,width=150); self.amo_monto.grid(row=0,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Tasa anual (%):").grid(row=1,column=0,padx=10,pady=5,sticky="e")
        self.amo_tasa = ctk.CTkEntry(frame,width=150); self.amo_tasa.grid(row=1,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Plazo (años):").grid(row=2,column=0,padx=10,pady=5,sticky="e")
        self.amo_plazo = ctk.CTkEntry(frame,width=150); self.amo_plazo.grid(row=2,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Periodicidad:").grid(row=3,column=0,padx=10,pady=5,sticky="e")
        self.amo_periodo = ctk.CTkOptionMenu(frame,values=["Mensual","Bimestral","Trimestral","Semestral","Anual"])
        self.amo_periodo.grid(row=3,column=1,padx=10,pady=5,sticky="w")
        btn_calc = ctk.CTkButton(frame,text="Calcular",command=self.calcular_amortizacion,fg_color=tema.success)
        btn_calc.grid(row=4,column=0,columnspan=2,pady=10)
        self.amo_resultado = ctk.CTkTextbox(frame,height=300,wrap="word")
        self.amo_resultado.grid(row=5,column=0,columnspan=2,padx=10,pady=10,sticky="ew")
        btn_reporte = ctk.CTkButton(frame,text="Generar reporte PDF",command=self.reporte_amortizacion,fg_color=tema.accent)
        btn_reporte.grid(row=6,column=0,columnspan=2,pady=5)

    def calcular_amortizacion(self):
        try:
            monto = float(self.amo_monto.get())
            tasa_anual = float(self.amo_tasa.get())/100
            años = int(self.amo_plazo.get())
            periodo = self.amo_periodo.get()
            if periodo == "Mensual": m=12
            elif periodo == "Bimestral": m=6
            elif periodo == "Trimestral": m=4
            elif periodo == "Semestral": m=2
            else: m=1
            tasa_periodo = tasa_anual / m
            n_periodos = años * m
            factor = (1 - (1+tasa_periodo)**-n_periodos)/tasa_periodo
            cuota = monto / factor
            saldo = monto
            tabla = []
            total_intereses = 0
            for i in range(1, n_periodos+1):
                interes = saldo * tasa_periodo
                amortizacion = cuota - interes
                saldo -= amortizacion
                total_intereses += interes
                tabla.append((i, interes, amortizacion, cuota, max(saldo,0)))
            texto = f"Cuota periódica: {cuota:,.2f}\nTotal intereses: {total_intereses:,.2f}\nTotal pagado: {cuota*n_periodos:,.2f}\n\nTabla de amortización (primeros 10 períodos):\n"
            texto += "Período   Interés   Amortización   Cuota   Saldo\n"
            for row in tabla[:10]:
                texto += f"{row[0]:8d} {row[1]:10,.2f} {row[2]:14,.2f} {row[3]:8,.2f} {row[4]:10,.2f}\n"
            if n_periodos>10: texto += "...\n"
            self.amo_resultado.delete("0.0","end")
            self.amo_resultado.insert("0.0", texto)
            self.amo_ultimo = ({"Monto":monto, "Tasa anual %":self.amo_tasa.get(), "Plazo años":años, "Periodicidad":periodo}, cuota, texto)
        except Exception as e:
            self.amo_resultado.delete("0.0","end")
            self.amo_resultado.insert("0.0",f"Error: {str(e)}")

    def reporte_amortizacion(self):
        if not hasattr(self, 'amo_ultimo'):
            messagebox.showwarning("Sin datos", "Primero calcule una amortización")
            return
        datos, cuota, tabla = self.amo_ultimo
        lineas = tabla.split('\n')
        saldos = []
        for linea in lineas:
            if 'Período' in linea or not linea.strip() or '...' in linea:
                continue
            partes = linea.split()
            if len(partes) >= 5:
                try:
                    saldo_str = partes[4].replace(',','')
                    saldo = float(saldo_str)
                    saldos.append(saldo)
                except:
                    pass
        if saldos:
            fig, ax = plt.subplots(figsize=(5,3), dpi=90)
            ax.plot(range(1,len(saldos)+1), saldos, marker='o', color=tema.accent)
            ax.set_title("Evolución del saldo del préstamo")
            ax.set_xlabel("Período")
            ax.set_ylabel("Saldo ($)")
            fig.tight_layout()
            grafico = "temp_amo.png"
            fig.savefig(grafico)
            plt.close(fig)
        else:
            grafico = None
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if archivo:
            analisis = f"Se calculó la tabla de amortización para un préstamo de ${datos['Monto']:,.2f} a una tasa anual del {datos['Tasa anual %']}%, con un plazo de {datos['Plazo años']} años y periodicidad {datos['Periodicidad']}. La cuota periódica es de ${cuota:,.2f}. El total de intereses pagados se muestra en la tabla. El saldo se reduce periódicamente hasta cero."
            generar_reporte_calculadora(datos, "AMORTIZACIÓN DE PRÉSTAMOS", analisis, grafico, archivo)
            messagebox.showinfo("Reporte", f"PDF guardado en {archivo}")

    # ---------- Conversión de Tasas ----------
    def build_sub_conversion_tasas(self,parent):
        frame = ctk.CTkFrame(parent,fg_color=tema.card,corner_radius=10)
        frame.pack(fill='both',expand=True,padx=10,pady=10)
        ctk.CTkLabel(frame,text="Tasa nominal anual (%):").grid(row=0,column=0,padx=10,pady=5,sticky="e")
        self.conv_nom = ctk.CTkEntry(frame,width=150); self.conv_nom.grid(row=0,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Períodos de capitalización por año:").grid(row=1,column=0,padx=10,pady=5,sticky="e")
        self.conv_m = ctk.CTkEntry(frame,width=150); self.conv_m.grid(row=1,column=1,padx=10,pady=5)
        ctk.CTkButton(frame,text="Calcular EAR",command=self.calcular_ear,fg_color=tema.accent).grid(row=2,column=0,columnspan=2,pady=5)
        self.conv_ear = ctk.CTkTextbox(frame,height=80,wrap="word")
        self.conv_ear.grid(row=3,column=0,columnspan=2,padx=10,pady=5,sticky="ew")
        ctk.CTkLabel(frame,text="Tasa efectiva anual (%):").grid(row=4,column=0,padx=10,pady=5,sticky="e")
        self.conv_efectiva = ctk.CTkEntry(frame,width=150); self.conv_efectiva.grid(row=4,column=1,padx=10,pady=5)
        ctk.CTkButton(frame,text="Calcular Tasa Nominal",command=self.calcular_nominal,fg_color=tema.accent).grid(row=5,column=0,columnspan=2,pady=5)
        self.conv_nom_res = ctk.CTkTextbox(frame,height=80,wrap="word")
        self.conv_nom_res.grid(row=6,column=0,columnspan=2,padx=10,pady=5,sticky="ew")
        btn_reporte = ctk.CTkButton(frame,text="Generar reporte PDF",command=self.reporte_conversion,fg_color=tema.accent)
        btn_reporte.grid(row=7,column=0,columnspan=2,pady=5)

    def calcular_ear(self):
        try:
            nominal = float(self.conv_nom.get())/100
            m = int(self.conv_m.get())
            ear = conversion_tasa_nominal_a_efectiva(nominal,m)*100
            self.conv_ear.delete("0.0","end")
            self.conv_ear.insert("0.0", f"Tasa Efectiva Anual = {ear:.4f}%")
            self.conv_ultimo = ({"Tasa nominal %": self.conv_nom.get(), "Capitalizaciones/año": m}, f"EAR = {ear:.4f}%", "La tasa efectiva anual se calcula a partir de la tasa nominal y la frecuencia de capitalización. Es la tasa real que se paga o recibe en un año.")
        except Exception as e:
            self.conv_ear.delete("0.0","end")
            self.conv_ear.insert("0.0",f"Error: {str(e)}")

    def calcular_nominal(self):
        try:
            efectiva = float(self.conv_efectiva.get())/100
            m = int(self.conv_m.get())
            nominal = conversion_tasa_efectiva_a_nominal(efectiva,m)*100
            self.conv_nom_res.delete("0.0","end")
            self.conv_nom_res.insert("0.0", f"Tasa Nominal Anual = {nominal:.4f}%")
            self.conv_ultimo = ({"Tasa efectiva %": self.conv_efectiva.get(), "Capitalizaciones/año": m}, f"Nominal = {nominal:.4f}%", "La tasa nominal anual se obtiene a partir de la tasa efectiva anual y la frecuencia de capitalización. Es la tasa que se declara en los contratos.")
        except Exception as e:
            self.conv_nom_res.delete("0.0","end")
            self.conv_nom_res.insert("0.0",f"Error: {str(e)}")

    def reporte_conversion(self):
        if not hasattr(self, 'conv_ultimo'):
            messagebox.showwarning("Sin datos", "Primero realice una conversión")
            return
        datos, res, analisis = self.conv_ultimo
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if archivo:
            generar_reporte_calculadora(datos, "CONVERSIÓN DE TASAS", analisis, None, archivo)
            messagebox.showinfo("Reporte", f"PDF guardado en {archivo}")

    # ---------- VPN / CAE / TIR (calculadora) ----------
    def build_sub_vpn_cae_tir(self,parent):
        frame = ctk.CTkFrame(parent,fg_color=tema.card,corner_radius=10)
        frame.pack(fill='both',expand=True,padx=10,pady=10)
        ctk.CTkLabel(frame,text="Método:").grid(row=0,column=0,padx=10,pady=5,sticky="e")
        self.vct_metodo = ctk.CTkOptionMenu(frame,values=["VPN","CAE","TIR"])
        self.vct_metodo.grid(row=0,column=1,padx=10,pady=5,sticky="w")
        self.vct_metodo.set("VPN")
        ctk.CTkLabel(frame,text="Inversión inicial:").grid(row=1,column=0,padx=10,pady=5,sticky="e")
        self.vct_inv = ctk.CTkEntry(frame,width=150); self.vct_inv.grid(row=1,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Flujos (separados por coma):").grid(row=2,column=0,padx=10,pady=5,sticky="e")
        self.vct_flujos = ctk.CTkEntry(frame,width=200); self.vct_flujos.grid(row=2,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Tasa %:").grid(row=3,column=0,padx=10,pady=5,sticky="e")
        self.vct_tasa = ctk.CTkEntry(frame,width=150); self.vct_tasa.grid(row=3,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Años:").grid(row=4,column=0,padx=10,pady=5,sticky="e")
        self.vct_anios = ctk.CTkEntry(frame,width=150); self.vct_anios.grid(row=4,column=1,padx=10,pady=5)
        ctk.CTkLabel(frame,text="Valor residual (opcional):").grid(row=5,column=0,padx=10,pady=5,sticky="e")
        self.vct_salv = ctk.CTkEntry(frame,width=150); self.vct_salv.grid(row=5,column=1,padx=10,pady=5)
        btn_calc = ctk.CTkButton(frame,text="Calcular",command=self.calcular_vpn_cae_tir,fg_color=tema.success)
        btn_calc.grid(row=6,column=0,columnspan=2,pady=10)
        self.vct_resultado = ctk.CTkTextbox(frame,height=150,wrap="word")
        self.vct_resultado.grid(row=7,column=0,columnspan=2,padx=10,pady=10,sticky="ew")
        btn_reporte = ctk.CTkButton(frame,text="Generar reporte PDF",command=self.reporte_vpn_cae_tir,fg_color=tema.accent)
        btn_reporte.grid(row=8,column=0,columnspan=2,pady=5)

    def calcular_vpn_cae_tir(self):
        try:
            metodo = self.vct_metodo.get()
            inversion = float(self.vct_inv.get()) if self.vct_inv.get() else 0
            flujos = [float(x.strip()) for x in self.vct_flujos.get().split(',') if x.strip()]
            tasa = float(self.vct_tasa.get())/100 if self.vct_tasa.get() else 0
            años = int(self.vct_anios.get()) if self.vct_anios.get() else len(flujos)
            salvage = float(self.vct_salv.get()) if self.vct_salv.get() else 0
            if len(flujos) < años:
                self.vct_resultado.delete("0.0","end")
                self.vct_resultado.insert("0.0","Error: número de flujos inferior a los años.")
                return
            flujos = flujos[:años]
            if metodo == "VPN":
                vpn, flujos_desc = calcular_vpn(inversion, flujos, tasa, años, salvage)
                analisis = f"El Valor Presente Neto es {vpn:,.2f}. " + ("Esto indica que el proyecto genera valor y es aceptable." if vpn>0 else "Esto indica que el proyecto destruye valor y no es aceptable.")
                self.vct_ultimo = ({"Inversión": inversion, "Flujos": flujos, "Tasa %": self.vct_tasa.get(), "Años": años, "Residual": salvage}, vpn, analisis, flujos)
                self.vct_resultado.delete("0.0","end")
                self.vct_resultado.insert("0.0",f"VPN = {vpn:,.2f}")
            elif metodo == "CAE":
                vpn, flujos_desc = calcular_vpn(inversion, flujos, tasa, años, salvage)
                cae = calcular_cae(vpn, tasa, años)
                analisis = f"El Costo Anual Equivalente es {cae:,.2f}. " + ("Negativo indica costos netos anuales." if cae<0 else "Positivo indica ingresos netos anuales.")
                self.vct_ultimo = ({"Inversión": inversion, "Flujos": flujos, "Tasa %": self.vct_tasa.get(), "Años": años, "Residual": salvage}, cae, analisis, flujos)
                self.vct_resultado.delete("0.0","end")
                self.vct_resultado.insert("0.0",f"CAE = {cae:,.2f}")
            else: # TIR
                tir = calcular_tir(inversion, flujos, años, salvage)
                if tir is None:
                    self.vct_resultado.delete("0.0","end")
                    self.vct_resultado.insert("0.0","No se pudo calcular TIR.")
                    return
                analisis = f"La Tasa Interna de Retorno es {tir*100:.4f}%. " + ("Supera la tasa de descuento, proyecto rentable." if tir>tasa else "No supera la tasa de descuento, proyecto no rentable.")
                self.vct_ultimo = ({"Inversión": inversion, "Flujos": flujos, "Años": años, "Residual": salvage}, tir, analisis, flujos)
                self.vct_resultado.delete("0.0","end")
                self.vct_resultado.insert("0.0",f"TIR = {tir*100:.4f}%")
        except Exception as e:
            self.vct_resultado.delete("0.0","end")
            self.vct_resultado.insert("0.0",f"Error: {str(e)}")

    def reporte_vpn_cae_tir(self):
        if not hasattr(self, 'vct_ultimo'):
            messagebox.showwarning("Sin datos", "Primero realice un cálculo")
            return
        datos, resultado, analisis, flujos = self.vct_ultimo
        fig, ax = plt.subplots(figsize=(5,3), dpi=90)
        ax.bar(range(1,len(flujos)+1), flujos, color=tema.accent)
        if "Inversión" in datos:
            inv = datos["Inversión"]
            ax.axhline(y=inv/len(flujos), color='red', linestyle='--', label='Inversión promedio/año')
        ax.set_title("Flujos de caja")
        ax.set_xlabel("Año")
        ax.set_ylabel("Monto ($)")
        ax.legend()
        fig.tight_layout()
        grafico = "temp_vct.png"
        fig.savefig(grafico)
        plt.close(fig)
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if archivo:
            titulo = f"CÁLCULO DE {self.vct_metodo.get()}"
            generar_reporte_calculadora(datos, titulo, analisis, grafico, archivo)
            messagebox.showinfo("Reporte", f"PDF guardado en {archivo}")

    # ================== PESTAÑA PROYECTOS ==================
    def build_proyectos_tab(self):
        self.proy_container = ctk.CTkScrollableFrame(self.tabview.tab("Proyectos"),fg_color=tema.bg)
        self.proy_container.pack(fill='both',expand=True,padx=10,pady=10)
        ctk.CTkButton(self.proy_container,text="Nuevo Proyecto",command=self.agregar_proyecto,fg_color=tema.accent).pack(pady=10)
        self.cargar_proyectos_en_ui()

    def agregar_proyecto(self,datos_duplicado=None):
        nuevo = {
            "id":None,"nombre":f"Proyecto {len(self.proyectos)+1}","metodo":"VPN",
            "inversion":10000.0,"años":1,"tasa":0.12,"salvage":0.0,"flujos":[1000.0],
            "vpn":0.0,"tir":0.0,"cae":0.0,"payback":"N/A","riesgo":"Medio","recomendacion":"",
            "flujo_entries":[],"flujos_container":None,"after_id":None
        }
        if datos_duplicado:
            nuevo.update(datos_duplicado)
            nuevo["nombre"] = datos_duplicado["nombre"]+" (copia)"
            nuevo["id"]=None; nuevo["flujos"]=datos_duplicado["flujos"][:]
        self.proyectos.append(nuevo)
        self.cargar_proyectos_en_ui()
        self.guardar_y_actualizar()

    def cargar_proyectos_en_ui(self):
        for f in self.project_frames: f.destroy()
        self.project_frames.clear()
        for idx,proy in enumerate(self.proyectos):
            frame = ctk.CTkFrame(self.proy_container,fg_color=tema.card,corner_radius=12,border_width=1,border_color=tema.border)
            frame.pack(fill='x',padx=15,pady=8)
            self.project_frames.append(frame)
            self._crear_editor_proyecto(frame,idx,proy)

    def _crear_editor_proyecto(self,parent,idx,proy):
        top = ctk.CTkFrame(parent,fg_color="transparent")
        top.pack(fill='x',padx=10,pady=5)
        entry_nombre = ctk.CTkEntry(top,width=200,font=("Segoe UI",14,"bold"))
        entry_nombre.insert(0,proy["nombre"])
        entry_nombre.pack(side='left',padx=5)
        def act_nombre():
            proy["nombre"]=entry_nombre.get()
            self.actualizar_dashboard()
        entry_nombre.bind('<KeyRelease>',lambda e: act_nombre())
        metodo_options=["VPN","CAE","TIR"]
        metodo_var = ctk.StringVar(value=proy.get("metodo","VPN"))
        metodo_menu = ctk.CTkOptionMenu(top,values=metodo_options,variable=metodo_var,width=100)
        metodo_menu.pack(side='left',padx=10)
        def act_metodo(*args):
            proy["metodo"]=metodo_var.get()
        metodo_var.trace_add('write',act_metodo)
        badge = ctk.CTkLabel(top, text="", width=60, height=25, corner_radius=5,
                             fg_color=tema.accent, text_color="white", font=("Segoe UI", 10, "bold"))
        badge.pack(side='left', padx=5)
        def actualizar_badge(*args):
            badge.configure(text=metodo_var.get())
        metodo_var.trace_add('write', actualizar_badge)
        actualizar_badge()
        btn_duplicar = ctk.CTkButton(top,text="Duplicar",command=lambda: self.agregar_proyecto(proy),fg_color=tema.accent,width=80)
        btn_duplicar.pack(side='right',padx=5)
        btn_eliminar = ctk.CTkButton(top,text="Eliminar",command=lambda: self.eliminar_proyecto(idx),fg_color=tema.danger,width=80)
        btn_eliminar.pack(side='right',padx=5)
        btn_limpiar = ctk.CTkButton(top,text="Limpiar",command=lambda: self.limpiar_proyecto(idx),fg_color=tema.warning,width=80)
        btn_limpiar.pack(side='right',padx=5)

        campos_frame = ctk.CTkFrame(parent,fg_color="transparent")
        campos_frame.pack(fill='x',padx=10,pady=5)
        campos_frame.grid_columnconfigure(1,weight=1)
        inv_entry = ctk.CTkEntry(campos_frame,width=150)
        inv_entry.insert(0,str(proy["inversion"]))
        inv_entry.grid(row=0,column=1,padx=5,pady=2,sticky="w")
        ctk.CTkLabel(campos_frame,text="Inversión inicial ($):").grid(row=0,column=0,padx=5,pady=2,sticky="e")
        proy["entry_inv"]=inv_entry
        anios_entry = ctk.CTkEntry(campos_frame,width=80)
        anios_entry.insert(0,str(proy["años"]))
        anios_entry.grid(row=1,column=1,padx=5,pady=2,sticky="w")
        ctk.CTkLabel(campos_frame,text="Vida útil (años):").grid(row=1,column=0,padx=5,pady=2,sticky="e")
        proy["entry_anios"]=anios_entry
        tasa_entry = ctk.CTkEntry(campos_frame,width=100)
        tasa_entry.insert(0,str(proy["tasa"]*100))
        tasa_entry.grid(row=2,column=1,padx=5,pady=2,sticky="w")
        ctk.CTkLabel(campos_frame,text="TMAR (%):").grid(row=2,column=0,padx=5,pady=2,sticky="e")
        proy["entry_tasa"]=tasa_entry
        salvage_entry = ctk.CTkEntry(campos_frame,width=150)
        salvage_entry.insert(0,str(proy["salvage"]))
        salvage_entry.grid(row=3,column=1,padx=5,pady=2,sticky="w")
        ctk.CTkLabel(campos_frame,text="Valor residual ($):").grid(row=3,column=0,padx=5,pady=2,sticky="e")
        proy["entry_salv"] = salvage_entry

        flujos_label = ctk.CTkFrame(parent, fg_color="transparent")
        flujos_label.pack(fill='x', padx=10, pady=(5,0))
        ctk.CTkLabel(flujos_label, text="Flujos anuales:", font=("Segoe UI", 12, "bold")).pack(anchor='w')
        flujos_scroll = ctk.CTkScrollableFrame(parent, orientation="horizontal", height=80, fg_color=tema.bg, border_width=1, border_color=tema.border)
        flujos_scroll.pack(fill='x', padx=10, pady=5)
        proy["flujos_container"] = flujos_scroll
        self.recargar_flujos_ui(proy)

        btn_add = ctk.CTkButton(parent, text="Agregar año", command=lambda: self.agregar_flujo_proyecto(proy), fg_color=tema.accent, width=140)
        btn_add.pack(pady=5)
        btn_calcular = ctk.CTkButton(parent, text=f"Calcular {proy['nombre']}", command=lambda: self.calcular_proyecto(idx), fg_color=tema.success)
        btn_calcular.pack(pady=5)
        warn_label = ctk.CTkLabel(parent, text="", font=("Segoe UI", 10), text_color=tema.warning)
        warn_label.pack(pady=2)
        proy["warn_label"] = warn_label
        res_label = ctk.CTkLabel(parent, text="", font=("Segoe UI", 11), text_color=tema.text_secondary)
        res_label.pack(pady=2)
        proy["res_label"] = res_label
        action_frame = ctk.CTkFrame(parent, fg_color="transparent")
        action_frame.pack(pady=5)
        ctk.CTkButton(action_frame, text="Copiar análisis", command=lambda: self.copiar_analisis(proy), width=110).pack(side='left', padx=2)
        ctk.CTkButton(action_frame, text="PDF", command=lambda: self.generar_reporte_pdf(idx), width=80).pack(side='left', padx=2)
        ctk.CTkButton(action_frame, text="Excel", command=lambda: self.generar_reporte_excel(idx), width=80).pack(side='left', padx=2)
        ctk.CTkButton(action_frame, text="Auditor", command=lambda: self.modo_auditor(proy), width=80, fg_color=tema.accent).pack(side='left', padx=2)
        ctk.CTkButton(action_frame, text="Boxplot", command=lambda: self.mostrar_boxplot(proy), width=80, fg_color=tema.accent).pack(side='left', padx=2)

    def recargar_flujos_ui(self, proy):
        for child in proy["flujos_container"].winfo_children():
            child.destroy()
        proy["flujo_entries"].clear()
        for i, val in enumerate(proy["flujos"]):
            frame_f = ctk.CTkFrame(proy["flujos_container"], fg_color="transparent")
            frame_f.pack(side='left', padx=5, pady=2)
            lbl = ctk.CTkLabel(frame_f, text=f"Año {i+1}:", font=("Segoe UI", 11))
            lbl.pack()
            entry = ctk.CTkEntry(frame_f, width=80)
            entry.insert(0, str(val))
            entry.pack()
            def autoguardar(e, p=proy):
                if p.get("after_id"):
                    self.after_cancel(p["after_id"])
                p["after_id"] = self.after(800, lambda: self.guardar_flujos_desde_ui(p))
            entry.bind('<KeyRelease>', autoguardar)
            entry.bind('<FocusOut>', lambda e, p=proy: self.guardar_flujos_desde_ui(p))
            proy["flujo_entries"].append(entry)

    def agregar_flujo_proyecto(self, proy):
        proy["flujos"].append(0.0)
        frame_f = ctk.CTkFrame(proy["flujos_container"], fg_color="transparent")
        frame_f.pack(side='left', padx=5, pady=2)
        lbl = ctk.CTkLabel(frame_f, text=f"Año {len(proy['flujos'])}:", font=("Segoe UI", 11))
        lbl.pack()
        entry = ctk.CTkEntry(frame_f, width=80)
        entry.insert(0, "0")
        entry.pack()
        entry.bind('<KeyRelease>', lambda e, p=proy: self.guardar_flujos_desde_ui(p))
        entry.bind('<FocusOut>', lambda e, p=proy: self.guardar_flujos_desde_ui(p))
        proy["flujo_entries"].append(entry)
        self.guardar_flujos_desde_ui(proy)

    def guardar_flujos_desde_ui(self, proy):
        nuevos = [float(e.get()) if e.get() else 0.0 for e in proy["flujo_entries"]]
        proy["flujos"] = nuevos
        if proy.get("after_id"):
            self.after_cancel(proy["after_id"])
            proy["after_id"] = None

    def limpiar_proyecto(self, idx):
        proy = self.proyectos[idx]
        if messagebox.askyesno("Limpiar", f"¿Reiniciar todos los datos de {proy['nombre']}?"):
            proy["inversion"] = 0.0
            proy["años"] = 1
            proy["tasa"] = 0.12
            proy["salvage"] = 0.0
            proy["flujos"] = [0.0]
            proy["vpn"] = 0.0
            proy["tir"] = 0.0
            proy["cae"] = 0.0
            proy["payback"] = "N/A"
            proy["riesgo"] = "Medio"
            proy["recomendacion"] = ""
            if proy["res_label"]:
                proy["res_label"].configure(text="")
                proy["warn_label"].configure(text="")
            self.cargar_proyectos_en_ui()
            self.guardar_y_actualizar()

    def copiar_analisis(self, proy):
        otros = [p for p in self.proyectos if p["nombre"] != proy["nombre"]]
        texto = analisis_profundo_para_reporte(proy, otros)
        self.clipboard_clear()
        self.clipboard_append(texto)
        self.status_bar.configure(text="Análisis copiado al portapapeles")

    def obtener_flujos_desde_entries(self, proy):
        flujos = []
        for entry in proy["flujo_entries"]:
            try:
                flujos.append(float(entry.get()))
            except:
                flujos.append(0.0)
        return flujos

    def calcular_proyecto(self, idx):
        p = self.proyectos[idx]
        inversion = float(p["entry_inv"].get())
        años = int(p["entry_anios"].get())
        tasa = float(p["entry_tasa"].get()) / 100.0
        salvage = float(p["entry_salv"].get()) if p["entry_salv"].get() else 0.0
        flujos = self.obtener_flujos_desde_entries(p)
        if len(flujos) < años:
            messagebox.showerror("Error", f"Se requieren {años} flujos")
            return
        metodo = p.get("metodo", "VPN")
        vpn, tir, cae = 0.0, 0.0, 0.0
        payback = calcular_payback_preciso(inversion, flujos)
        if metodo == "VPN":
            vpn, _ = calcular_vpn(inversion, flujos, tasa, años, salvage)
            tir = calcular_tir(inversion, flujos, años, salvage) or 0.0
            cae = calcular_cae(vpn, tasa, años)
        elif metodo == "CAE":
            vpn_temp, _ = calcular_vpn(inversion, flujos, tasa, años, salvage)
            cae = calcular_cae(vpn_temp, tasa, años)
            tir = calcular_tir(inversion, flujos, años, salvage) or 0.0
            vpn = vpn_temp
        else:
            tir = calcular_tir(inversion, flujos, años, salvage) or 0.0
            vpn, _ = calcular_vpn(inversion, flujos, tasa, años, salvage)
            cae = calcular_cae(vpn, tasa, años)
        p.update({
            "inversion": inversion, "años": años, "tasa": tasa,
            "salvage": salvage, "flujos": flujos,
            "vpn": vpn, "tir": tir, "cae": cae,
            "payback": f"{payback:.2f} años" if payback else "> vida útil"
        })
        p["riesgo"] = clasificar_riesgo(p)
        p["recomendacion"] = analisis_profundo_para_reporte(p, [pr for pr in self.proyectos if pr["nombre"] != p["nombre"]])
        advertencias = validaciones_inteligentes(p)
        p["warn_label"].configure(text=" | ".join(advertencias) if advertencias else "")
        if metodo == "VPN":
            principal = f"VPN: ${vpn:,.2f}"
            secundario = f" | TIR: {tir*100:.2f}% | CAE: ${cae:,.2f}"
        elif metodo == "CAE":
            principal = f"CAE: ${cae:,.2f} anual"
            secundario = f" | VPN: ${vpn:,.2f} | TIR: {tir*100:.2f}%"
        else:
            principal = f"TIR: {tir*100:.2f}%"
            secundario = f" | VPN: ${vpn:,.2f} | CAE: ${cae:,.2f}"
        p["res_label"].configure(text=principal + secundario)
        guardar_proyecto(p)
        self.guardar_y_actualizar()
        self.status_bar.configure(text=f"Proyecto {p['nombre']} calculado y guardado.")

    def eliminar_proyecto(self, idx):
        proy = self.proyectos[idx]
        if messagebox.askyesno("Confirmar", f"¿Eliminar permanentemente '{proy['nombre']}'?"):
            if proy.get("id"):
                eliminar_proyecto(proy["id"])
            self.proyectos.pop(idx)
            self.cargar_proyectos_en_ui()
            self.guardar_y_actualizar()

    def generar_reporte_pdf(self, idx):
        p = self.proyectos[idx]
        if p["vpn"] == 0 and p["tir"] == 0 and p["cae"] == 0:
            messagebox.showwarning("Sin datos", "Primero calcule el proyecto")
            return
        otros = [proy for proy in self.proyectos if proy["nombre"] != p["nombre"]]
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if archivo:
            generar_reporte_pdf_avanzado(p, otros, archivo)
            messagebox.showinfo("Reporte", f"PDF generado: {archivo}")

    def generar_reporte_excel(self, idx):
        p = self.proyectos[idx]
        if p["vpn"] == 0 and p["tir"] == 0 and p["cae"] == 0:
            messagebox.showwarning("Sin datos", "Primero calcule el proyecto")
            return
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if archivo:
            generar_reporte_excel(p, archivo)
            messagebox.showinfo("Reporte", f"Excel generado: {archivo}")

    def modo_auditor(self, proyecto):
        ventana = ctk.CTkToplevel(self)
        ventana.title(f"Auditor - {proyecto['nombre']}")
        ventana.geometry("700x500")
        ventana.configure(fg_color=tema.bg)
        texto = ctk.CTkTextbox(ventana, wrap="word", font=("Consolas", 11))
        texto.pack(fill='both', expand=True, padx=10, pady=10)
        inv = proyecto["inversion"]
        tasa = proyecto["tasa"]
        años = proyecto["años"]
        salvage = proyecto["salvage"]
        flujos = proyecto["flujos"][:años]
        vpn, flujos_desc = calcular_vpn(inv, flujos, tasa, años, salvage)
        contenido = f"PROYECTO: {proyecto['nombre']}\nMétodo: {proyecto.get('metodo','VPN')}\nInversión: ${inv:,.2f}\nTMAR: {tasa*100:.2f}%\nVida útil: {años} años\nValor residual: ${salvage:,.2f}\n\nCÁLCULO VPN:\nVPN = -${inv:,.2f}"
        for t, (f, fd) in enumerate(zip(flujos, flujos_desc), 1):
            contenido += f" + {f:,.2f} / (1+{tasa:.4f})^{t} = {fd:,.2f}\n"
        contenido += f"VPN final = ${vpn:,.2f}\n\nTIR = {proyecto['tir']*100:.4f}%\n\nPayback:\n"
        acum = 0
        for i, f in enumerate(flujos, 1):
            acum += f
            contenido += f"Año {i}: ${acum:,.2f}\n"
            if acum >= inv:
                contenido += f"Recuperación en año {i}\n"
                break
        texto.insert("0.0", contenido)
        texto.configure(state="disabled")

    def mostrar_boxplot(self, proyecto):
        flujos = proyecto["flujos"]
        if not flujos:
            messagebox.showinfo("Boxplot", "No hay flujos para graficar.")
            return
        fig, ax = plt.subplots(figsize=(6, 4), facecolor=tema.card)
        ax.boxplot(flujos, patch_artist=True, boxprops=dict(facecolor=tema.accent))
        ax.set_title(f"Distribución de Flujos - {proyecto['nombre']}", color=tema.text_primary)
        ax.tick_params(colors=tema.text_secondary)
        ventana = ctk.CTkToplevel(self)
        ventana.title(f"Boxplot - {proyecto['nombre']}")
        ventana.geometry("600x400")
        canvas = FigureCanvasTkAgg(fig, master=ventana)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    # ================== DASHBOARD ==================
    def build_dashboard_tab(self):
        self.dash_container = ctk.CTkScrollableFrame(self.tabview.tab("Dashboard"), fg_color=tema.bg)
        self.dash_container.pack(fill='both', expand=True, padx=10, pady=10)
        filtro_frame = ctk.CTkFrame(self.dash_container, fg_color="transparent")
        filtro_frame.pack(fill='x', pady=5)
        ctk.CTkLabel(filtro_frame, text="Filtrar por método:", font=("Segoe UI", 12)).pack(side='left', padx=5)
        self.dash_filtro = ctk.CTkOptionMenu(filtro_frame, values=["Todos","VPN","CAE","TIR"],
                                             command=lambda _: self.actualizar_dashboard())
        self.dash_filtro.pack(side='left', padx=10)
        self.dash_filtro.set("Todos")
        self.dash_contenido = ctk.CTkFrame(self.dash_container, fg_color="transparent")
        self.dash_contenido.pack(fill='both', expand=True)
        self.actualizar_dashboard()

    def actualizar_dashboard(self):
        for w in self.dash_contenido.winfo_children():
            w.destroy()
        filtro = self.dash_filtro.get()
        proyectos_filtrados = [p for p in self.proyectos if filtro == "Todos" or p.get("metodo") == filtro]
        if not proyectos_filtrados:
            ctk.CTkLabel(self.dash_contenido, text="No hay proyectos con el método seleccionado.",
                         font=("Segoe UI", 14), text_color=tema.text_secondary).pack(pady=30)
            return
        total = len(proyectos_filtrados)
        aceptan = sum(1 for p in proyectos_filtrados if decidir_aceptacion(p))
        rechazan = total - aceptan
        vpn_avg = np.mean([p['vpn'] for p in proyectos_filtrados]) if proyectos_filtrados else 0
        tir_avg = np.mean([p['tir']*100 for p in proyectos_filtrados]) if proyectos_filtrados else 0
        cae_avg = np.mean([p['cae'] for p in proyectos_filtrados]) if proyectos_filtrados else 0

        tarjetas_frame = ctk.CTkFrame(self.dash_contenido, fg_color="transparent")
        tarjetas_frame.pack(fill='x', pady=10)
        self._crear_tarjeta_resumen(tarjetas_frame, "Total Proyectos", str(total), 0, 0)
        self._crear_tarjeta_resumen(tarjetas_frame, "Aceptados", str(aceptan), 0, 1, tema.success)
        self._crear_tarjeta_resumen(tarjetas_frame, "Rechazados", str(rechazan), 0, 2, tema.danger)
        riesgos = [1 if p['riesgo']=='Bajo' else 2 if p['riesgo']=='Medio' else 3 for p in proyectos_filtrados]
        riesgo_prom = np.mean(riesgos) if riesgos else 0
        self._crear_tarjeta_resumen(tarjetas_frame, "Riesgo Promedio", f"{riesgo_prom:.1f}/3", 0, 3)
        self._crear_tarjeta_resumen(tarjetas_frame, "VPN Promedio", f"${vpn_avg:,.2f}", 1, 0)
        self._crear_tarjeta_resumen(tarjetas_frame, "TIR Promedio", f"{tir_avg:.2f}%", 1, 1)
        self._crear_tarjeta_resumen(tarjetas_frame, "CAE Promedio", f"${cae_avg:,.2f}", 1, 2)
        self._crear_tarjeta_resumen(tarjetas_frame, "Método", filtro, 1, 3, tema.accent)

        fig = plt.figure(figsize=(12, 5), facecolor=tema.card)
        ax1 = fig.add_subplot(1, 2, 1, facecolor=tema.card)
        nombres = [p['nombre'] for p in proyectos_filtrados]
        vpns = [p['vpn'] for p in proyectos_filtrados]
        colores = [tema.success if v > 0 else tema.danger for v in vpns]
        ax1.bar(nombres, vpns, color=colores, edgecolor='white')
        ax1.axhline(y=0, color='white', linestyle='--', linewidth=1)
        for i, v in enumerate(vpns):
            ax1.text(i, v, f"${v:,.0f}", ha='center', va='bottom' if v>0 else 'top', fontsize=8, color=tema.text_primary)
        ax1.set_title("VPN por Proyecto", color=tema.text_primary)
        ax1.tick_params(colors=tema.text_secondary, labelsize=8)
        ax2 = fig.add_subplot(1, 2, 2, facecolor=tema.card)
        if total>0:
            ax2.pie([aceptan, rechazan], labels=["Aceptar","Rechazar"], autopct='%1.1f%%',
                    startangle=90, colors=[tema.success, tema.danger],
                    wedgeprops=dict(width=0.4, edgecolor='white'))
            centre = plt.Circle((0,0),0.7,fc=tema.card, edgecolor='white', linewidth=1)
            ax2.add_artist(centre)
            ax2.text(0,0,f"{total}\nproyectos",ha='center',va='center',fontsize=12,weight='bold',color=tema.text_primary)
        ax2.set_title("Decisiones", color=tema.text_primary)
        canvas = FigureCanvasTkAgg(fig, master=self.dash_contenido)
        canvas.draw()
        canvas.get_tk_widget().pack(pady=10)

        ctk.CTkLabel(self.dash_contenido, text="Proyectos Destacados", font=("Segoe UI", 14, "bold")).pack(anchor='w', padx=20, pady=(10,0))
        tree = ttk.Treeview(self.dash_contenido, columns=("nombre","metodo","vpn","tir","cae","riesgo","decision"), show='headings', height=min(5, total))
        for col in ("nombre","metodo","vpn","tir","cae","riesgo","decision"):
            tree.heading(col, text=col.capitalize())
            tree.column(col, width=80 if col!="nombre" else 120)
        for p in sorted(proyectos_filtrados, key=lambda x: x['vpn'], reverse=True)[:5]:
            dec = "Aceptar" if decidir_aceptacion(p) else "Rechazar"
            tree.insert("", "end", values=(p["nombre"], p.get("metodo"), f"${p['vpn']:,.0f}", f"{p['tir']*100:.1f}%", f"${p['cae']:,.0f}", p["riesgo"], dec))
        tree.pack(padx=20, pady=10, fill='x')

    def _crear_tarjeta_resumen(self, parent, titulo, valor, fila, col, color=None):
        frame = ctk.CTkFrame(parent, fg_color=tema.card, corner_radius=10)
        frame.grid(row=fila, column=col, padx=8, pady=5, sticky="nsew")
        parent.grid_columnconfigure(col, weight=1)
        ctk.CTkLabel(frame, text=titulo, font=("Segoe UI", 10), text_color=tema.text_secondary).pack(pady=(8,0))
        lbl_valor = ctk.CTkLabel(frame, text=valor, font=("Segoe UI", 18, "bold"), text_color=color if color else tema.text_primary)
        lbl_valor.pack(pady=(0,8))

    # ================== GRÁFICOS ==================
    def build_graficos_tab(self):
        self.graf_container = ctk.CTkScrollableFrame(self.tabview.tab("Gráficos"), fg_color=tema.bg)
        self.graf_container.pack(fill='both', expand=True, padx=10, pady=10)
        self.actualizar_graficos_tab()

    def actualizar_graficos_tab(self):
        for w in self.graf_container.winfo_children():
            w.destroy()
        if not self.proyectos:
            return
        fig1, ax1 = plt.subplots(figsize=(8,4), dpi=90, facecolor=tema.card)
        for p in self.proyectos:
            if p.get("vpn",0)!=0:
                _, flujos_desc = calcular_vpn(p["inversion"],p["flujos"],p["tasa"],p["años"],p["salvage"])
                acum = np.cumsum(flujos_desc)
                ax1.plot(range(1,len(acum)+1), acum, marker='o', label=f"{p['nombre']} ({p.get('metodo','VPN')})")
        ax1.axhline(y=0, color='white', linestyle='--')
        ax1.set_title("Evolución del VPN Acumulado por Proyecto", color=tema.text_primary)
        ax1.tick_params(colors=tema.text_secondary)
        ax1.legend(facecolor=tema.card, labelcolor=tema.text_primary, fontsize=8)
        canvas1 = FigureCanvasTkAgg(fig1, master=self.graf_container)
        canvas1.draw()
        canvas1.get_tk_widget().pack(pady=10)
        if HOVER_AVAILABLE:
            mplcursors.cursor(canvas1, hover=True).connect("add", lambda sel: sel.annotation.set_text(f"{sel.artist.get_label()}\nAño {int(sel.target[0])}\n${sel.target[1]:,.2f}"))
        self._crear_radar()

    def _crear_radar(self):
        if len(self.proyectos)==0: return
        indicadores = ['VPN','TIR','CAE','Payback']
        def norm(val, tipo, max_val):
            if tipo in ('CAE','Payback'):
                return 1 - (val/max_val) if max_val!=0 else 0
            else:
                return val/max_val if max_val!=0 else 0
        max_vpn = max(p["vpn"] for p in self.proyectos)
        max_tir = max(p["tir"] for p in self.proyectos)
        max_cae = max(p["cae"] for p in self.proyectos)
        max_pb = max(float(p["payback"].split()[0]) if p["payback"].replace('.','',1).isdigit() else 1 for p in self.proyectos)
        angulos = np.linspace(0,2*np.pi,len(indicadores),endpoint=False).tolist()
        angulos += angulos[:1]
        fig, ax = plt.subplots(figsize=(5,4), subplot_kw={'projection':'polar'}, facecolor=tema.card)
        for p in self.proyectos:
            vals = [norm(p["vpn"],'VPN',max_vpn), norm(p["tir"],'TIR',max_tir),
                    norm(p["cae"],'CAE',max_cae), norm(float(p["payback"].split()[0]) if p["payback"].replace('.','',1).isdigit() else 1,'Payback',max_pb)]
            vals += vals[:1]
            ax.plot(angulos, vals, 'o-', linewidth=2, label=p["nombre"])
            ax.fill(angulos, vals, alpha=0.25)
        ax.set_xticks(angulos[:-1])
        ax.set_xticklabels(indicadores, color=tema.text_primary)
        ax.legend(loc='upper right', bbox_to_anchor=(1.3,1.0), facecolor=tema.card, labelcolor=tema.text_primary)
        canvas = FigureCanvasTkAgg(fig, master=self.graf_container)
        canvas.draw()
        canvas.get_tk_widget().pack(pady=10)

    # ================== COMPARACIÓN ==================
    def build_comparacion_tab(self):
        self.comp_container = ctk.CTkScrollableFrame(self.tabview.tab("Comparación"), fg_color=tema.bg)
        self.comp_container.pack(fill='both', expand=True, padx=10, pady=10)
        self.actualizar_comparacion_tab()

    def actualizar_comparacion_tab(self):
        for w in self.comp_container.winfo_children():
            w.destroy()
        if len(self.proyectos)<2:
            ctk.CTkLabel(self.comp_container, text="Se necesitan al menos 2 proyectos para comparar.").pack()
            return
        ctk.CTkLabel(self.comp_container, text="Tabla de proyectos", font=("Segoe UI", 14, "bold")).pack(anchor='w', padx=10)
        tree = ttk.Treeview(self.comp_container, columns=("Proyecto","Método","VPN","TIR","CAE","Payback","Riesgo","Decisión"), show='headings')
        for col in ("Proyecto","Método","VPN","TIR","CAE","Payback","Riesgo","Decisión"):
            tree.heading(col, text=col)
            tree.column(col, width=90)
        for p in self.proyectos:
            decision = "Aceptar" if decidir_aceptacion(p) else "Rechazar"
            tree.insert("", 'end', values=(p["nombre"], p.get("metodo","VPN"), f"${p['vpn']:,.2f}", f"{p['tir']*100:.2f}%", f"${p['cae']:,.2f}", p["payback"], p["riesgo"], decision))
        tree.pack(fill='x', padx=10, pady=5)
        ctk.CTkButton(self.comp_container, text="Comparar dos proyectos (detalle)", command=self.comparar_dos_proyectos, fg_color=tema.accent).pack(pady=5)
        ctk.CTkButton(self.comp_container, text="Reporte Comparativo Global (PDF)", command=self.generar_reporte_comparativo, fg_color=tema.success).pack(pady=5)

    def comparar_dos_proyectos(self):
        if len(self.proyectos)<2:
            messagebox.showinfo("Info","Necesita al menos dos proyectos")
            return
        sel = self.proyectos[:2]
        ventana = ctk.CTkToplevel(self)
        ventana.title("Comparación Detallada")
        ventana.geometry("900x500")
        texto = ctk.CTkTextbox(ventana, wrap="word", font=("Consolas",11))
        texto.pack(fill='both', expand=True, padx=10, pady=10)
        comp = f"COMPARACIÓN: {sel[0]['nombre']} vs {sel[1]['nombre']}\n\n"
        for p in sel:
            comp += f"{p['nombre']} [{p.get('metodo','VPN')}]: VPN={p['vpn']:,.2f}, TIR={p['tir']*100:.2f}%, CAE={p['cae']:,.2f}, Payback={p['payback']}, Riesgo={p['riesgo']}\n"
        texto.insert("0.0", comp)
        texto.configure(state="disabled")

    def generar_reporte_comparativo(self):
        if len(self.proyectos) < 2:
            messagebox.showwarning("Info", "Se necesitan al menos dos proyectos.")
            return
        archivo = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if archivo:
            generar_reporte_comparativo_pdf(self.proyectos, archivo)
            messagebox.showinfo("Reporte", f"Reporte comparativo guardado en {archivo}")

    # ================== SENSIBILIDAD ==================
    def build_sensibilidad_tab(self):
        frame = ctk.CTkFrame(self.tabview.tab("Sensibilidad"), fg_color=tema.card, corner_radius=15)
        frame.pack(fill='both', expand=True, padx=15, pady=15)
        ctk.CTkLabel(frame, text="Análisis de Sensibilidad (VPN vs Tasa)", font=("Segoe UI", 16, "bold")).pack(pady=10)
        control_frame = ctk.CTkFrame(frame, fg_color="transparent")
        control_frame.pack(pady=5)
        ctk.CTkLabel(control_frame, text="Inversión ($):").grid(row=0, column=0, padx=5)
        self.sens_inv = ctk.CTkEntry(control_frame, width=120)
        self.sens_inv.grid(row=0, column=1, padx=5)
        self.sens_inv.insert(0, "10000")
        ctk.CTkLabel(control_frame, text="Años:").grid(row=0, column=2, padx=5)
        self.sens_anios = ctk.CTkEntry(control_frame, width=80)
        self.sens_anios.grid(row=0, column=3, padx=5)
        self.sens_anios.insert(0, "5")
        ctk.CTkLabel(control_frame, text="Tasa base (%):").grid(row=0, column=4, padx=5)
        self.sens_tasa = ctk.CTkEntry(control_frame, width=80)
        self.sens_tasa.grid(row=0, column=5, padx=5)
        self.sens_tasa.insert(0, "12")
        ctk.CTkLabel(control_frame, text="Flujos (separados por coma):").grid(row=1, column=0, columnspan=2, padx=5, pady=5)
        self.sens_flujos = ctk.CTkEntry(control_frame, width=300)
        self.sens_flujos.grid(row=1, column=2, columnspan=4, pady=5)
        self.sens_flujos.insert(0, "3000,3000,3000,3000,3000")
        ctk.CTkLabel(frame, text="Rango de variación (%):").pack(pady=5)
        self.rango_slider = ctk.CTkSlider(frame, from_=10, to=100, number_of_steps=9, width=400)
        self.rango_slider.set(50)
        self.rango_slider.pack(pady=5)
        btn_sens = ctk.CTkButton(frame, text="Generar Gráfico", command=self.graficar_sensibilidad, fg_color=tema.accent)
        btn_sens.pack(pady=10)
        self.fig_sens = plt.Figure(figsize=(8, 4), dpi=90, facecolor=tema.card)
        self.ax_sens = self.fig_sens.add_subplot(111, facecolor=tema.card)
        self.canvas_sens = FigureCanvasTkAgg(self.fig_sens, master=frame)
        self.canvas_sens.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)

    def graficar_sensibilidad(self):
        try:
            inv = float(self.sens_inv.get())
            años = int(self.sens_anios.get())
            tasa_base = float(self.sens_tasa.get()) / 100.0
            flujos = [float(x.strip()) for x in self.sens_flujos.get().split(',') if x.strip()]
            if len(flujos) < años:
                messagebox.showerror("Error", "Flujos insuficientes")
                return
            flujos = flujos[:años]
            rango = self.rango_slider.get()
            tasa_min = tasa_base * (1 - rango / 100)
            tasa_max = tasa_base * (1 + rango / 100)
            if tasa_min <= 0:
                tasa_min = 0.001
            step = (tasa_max - tasa_min) / 30
            tasas = [tasa_min + i * step for i in range(31)]
            vpns = [calcular_vpn(inv, flujos, t, años)[0] for t in tasas]
            self.fig_sens.clear()
            ax = self.fig_sens.add_subplot(111, facecolor=tema.card)
            ax.plot(tasas, vpns, color=tema.accent, linewidth=2)
            ax.axhline(y=0, color='red', linestyle='--')
            ax.fill_between(tasas, 0, vpns, where=(np.array(vpns) > 0), color=tema.success, alpha=0.3)
            ax.fill_between(tasas, 0, vpns, where=(np.array(vpns) < 0), color=tema.danger, alpha=0.3)
            ax.set_xlabel("Tasa de descuento", color=tema.text_secondary)
            ax.set_ylabel("VPN", color=tema.text_secondary)
            ax.set_title("Sensibilidad del VPN", color=tema.text_primary)
            ax.tick_params(colors=tema.text_secondary)
            self.canvas_sens.draw()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ================== HISTORIAL ==================
    def build_historial_tab(self):
        frame = ctk.CTkFrame(self.tabview.tab("Historial"), fg_color=tema.card, corner_radius=15)
        frame.pack(fill='both', expand=True, padx=15, pady=15)
        tree_frame = ctk.CTkFrame(frame, fg_color="transparent")
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)
        self.hist_tree = ttk.Treeview(tree_frame, columns=("id", "nombre", "metodo", "vpn", "tir", "cae", "fecha"), show='headings')
        self.hist_tree.heading("id", text="ID")
        self.hist_tree.heading("nombre", text="Proyecto")
        self.hist_tree.heading("metodo", text="Método")
        self.hist_tree.heading("vpn", text="VPN ($)")
        self.hist_tree.heading("tir", text="TIR (%)")
        self.hist_tree.heading("cae", text="CAE ($)")
        self.hist_tree.heading("fecha", text="Fecha")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.hist_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.hist_tree.xview)
        self.hist_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.hist_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(pady=10)
        ctk.CTkButton(btn_frame, text="Eliminar", command=self.eliminar_historial, fg_color=tema.danger).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="Exportar CSV", command=self.exportar_historial_csv).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="Comparar seleccionados", command=self.comparar_historicos, fg_color=tema.accent).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="Importar JSON", command=self.importar_proyectos_json).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="Exportar JSON", command=self.exportar_proyectos_json).pack(side='left', padx=5)
        self.cargar_historial_tabla()

    def cargar_historial_tabla(self):
        for item in self.hist_tree.get_children():
            self.hist_tree.delete(item)
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT id, nombre, metodo, vpn, tir, cae, fecha_creacion FROM proyectos ORDER BY fecha_creacion DESC")
        rows = c.fetchall()
        conn.close()
        for r in rows:
            self.hist_tree.insert("", 'end', values=(r[0], r[1], r[2], f"${r[3]:,.2f}", f"{r[4] * 100:.2f}%", f"${r[5]:,.2f}", r[6][:19]))

    def eliminar_historial(self):
        sel = self.hist_tree.selection()
        if sel:
            id_ = self.hist_tree.item(sel[0])['values'][0]
            if messagebox.askyesno("Confirmar", f"¿Eliminar proyecto ID {id_}?"):
                eliminar_proyecto(id_)
                self.cargar_historial_tabla()
                self.cargar_proyectos_guardados()
                self.cargar_proyectos_en_ui()
                self.guardar_y_actualizar()

    def exportar_historial_csv(self):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT id, nombre, metodo, vpn, tir, cae, fecha_creacion FROM proyectos")
        rows = c.fetchall()
        conn.close()
        if not rows:
            messagebox.showinfo("Vacío", "No hay proyectos en el historial.")
            return
        archivo = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if archivo:
            with open(archivo, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Nombre", "Método", "VPN", "TIR", "CAE", "Fecha"])
                for r in rows:
                    writer.writerow([r[0], r[1], r[2], r[3], r[4], r[5], r[6]])
            messagebox.showinfo("Exportado", f"Historial exportado a {archivo}")

    def comparar_historicos(self):
        sel = self.hist_tree.selection()
        if len(sel) != 2:
            messagebox.showwarning("Selección", "Seleccione exactamente dos proyectos del historial.")
            return
        ids = [self.hist_tree.item(s)['values'][0] for s in sel]
        proy1 = obtener_proyecto_por_id(ids[0])
        proy2 = obtener_proyecto_por_id(ids[1])
        if not proy1 or not proy2:
            return
        ventana = ctk.CTkToplevel(self)
        ventana.title("Comparación Histórica")
        ventana.geometry("800x500")
        texto = ctk.CTkTextbox(ventana, wrap="word", font=("Consolas", 11))
        texto.pack(fill='both', expand=True, padx=10, pady=10)
        comp = f"COMPARACIÓN: {proy1['nombre']} vs {proy2['nombre']}\n\n"
        comp += f"VPN: ${proy1['vpn']:,.2f} vs ${proy2['vpn']:,.2f}\n"
        comp += f"TIR: {proy1['tir'] * 100:.2f}% vs {proy2['tir'] * 100:.2f}%\n"
        comp += f"CAE: ${proy1['cae']:,.2f} vs ${proy2['cae']:,.2f}\n"
        comp += f"Riesgo: {proy1['riesgo']} vs {proy2['riesgo']}\n\n"
        mejor = proy1 if proy1['vpn'] > proy2['vpn'] else proy2
        comp += f"Mejor según VPN: {mejor['nombre']}\n{mejor['recomendacion']}"
        texto.insert("0.0", comp)
        texto.configure(state="disabled")

    def importar_proyectos_json(self):
        archivo = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if archivo:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    datos = json.load(f)
                for item in datos:
                    self.agregar_proyecto(item)
                self.guardar_y_actualizar()
                messagebox.showinfo("Importado", f"Se importaron {len(datos)} proyectos.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def exportar_proyectos_json(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
        if archivo:
            datos = [{"nombre": p["nombre"], "metodo": p.get("metodo", "VPN"), "inversion": p["inversion"], "flujos": p["flujos"],
                      "años": p["años"], "tasa": p["tasa"], "salvage": p["salvage"]} for p in self.proyectos]
            with open(archivo, 'w', encoding='utf-8') as f:
                json.dump(datos, f, indent=4, ensure_ascii=False)
            messagebox.showinfo("Exportado", f"Proyectos exportados a {archivo}")

    # ================== GLOSARIO ==================
    def build_glosario_tab(self):
        tab = self.tabview.tab("Glosario")
        frame = ctk.CTkFrame(tab, fg_color=tema.card, corner_radius=15)
        frame.pack(fill='both', expand=True, padx=15, pady=15)
        text_widget = tk.Text(frame, wrap="word", font=("Segoe UI", 11), bg=tema.card, fg=tema.text_primary,
                              bd=0, padx=10, pady=10, selectbackground=tema.accent)
        text_widget.pack(fill='both', expand=True)
        text_widget.tag_configure("titulo", font=("Segoe UI", 14, "bold"), foreground=tema.accent, spacing3=10)
        text_widget.tag_configure("subtitulo", font=("Segoe UI", 12, "bold"), foreground=tema.success, spacing2=5)
        glosario = """VPN (Valor Presente Neto)
Definición: Suma de los flujos de caja descontados menos la inversión inicial.
Fórmula: VPN = -Inversión + Σ (Flujo_t / (1 + i)^t) + Valor_residual / (1 + i)^n
Interpretación: VPN > 0 → proyecto rentable.

TIR (Tasa Interna de Retorno)
Tasa que hace VPN = 0. Si TIR > TMAR, aceptar.

CAE (Costo Anual Equivalente)
Anualidad uniforme. Fórmula: CAE = VPN × [ i(1+i)^n / ((1+i)^n - 1) ]
Útil para comparar proyectos de diferente vida útil.

TMAR: Tasa Mínima Atractiva de Retorno (costo de capital).
Payback: Tiempo para recuperar la inversión (simple).
Riesgo: Clasificación automática (Bajo/Medio/Alto) según volatilidad, margen TIR-TMAR y payback.
"""
        text_widget.insert("1.0", glosario, "titulo")
        text_widget.configure(state="disabled")

    # ================== AYUDA ==================
    def build_ayuda_tab(self):
        frame = ctk.CTkFrame(self.tabview.tab("Ayuda"), fg_color=tema.card, corner_radius=15)
        frame.pack(fill='both', expand=True, padx=15, pady=15)
        ctk.CTkLabel(frame, text="Asistente Virtual", font=("Segoe UI", 16, "bold")).pack(pady=10)
        ctk.CTkLabel(frame, text="Haz clic en una pregunta para obtener ayuda:", font=("Segoe UI", 12)).pack()
        preguntas = [
            ("Cómo seleccionar el método de evaluación", "En la pestaña Proyectos, use el menú desplegable junto al nombre del proyecto para elegir entre VPN, CAE o TIR."),
            ("Qué diferencia hay entre los métodos", "VPN mide valor en dinero actual, CAE anualiza los flujos, TIR da la rentabilidad porcentual."),
            ("Puedo combinar proyectos con diferentes métodos", "Sí, pero la comparación directa solo es posible entre proyectos del mismo método."),
            ("Qué significan las advertencias", "Indican posibles errores en los datos, como flujos insuficientes o tasas muy altas."),
            ("Cómo duplicar un proyecto", "En la pestaña Proyectos, cada tarjeta tiene un botón 'Duplicar'."),
            ("Puedo generar reportes en PDF con gráficos", "Sí, cada proyecto tiene botón PDF. El informe incluye gráficos y análisis detallado.")
        ]
        for pregunta, respuesta in preguntas:
            btn = ctk.CTkButton(frame, text=pregunta, command=lambda r=respuesta: self.mostrar_respuesta(r), fg_color="transparent", anchor="w")
            btn.pack(fill='x', padx=20, pady=2)
        self.ayuda_respuesta = ctk.CTkTextbox(frame, height=180, wrap="word", font=("Segoe UI", 11))
        self.ayuda_respuesta.pack(fill='x', padx=20, pady=10)

    def mostrar_respuesta(self, respuesta):
        self.ayuda_respuesta.delete("0.0", "end")
        self.ayuda_respuesta.insert("0.0", respuesta)

# ================== EJECUCIÓN ==================
if __name__ == "__main__":
    app = SEAE_Completo()
    app.mainloop()